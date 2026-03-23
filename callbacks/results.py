"""Sonuç Sekmesi — Model sonuçlarını detaylı accordion yapısında gösterir."""

import dash
from dash import dcc, html, Input, Output, State, dash_table, clientside_callback
import dash_bootstrap_components as dbc
import plotly.graph_objects as go

from app_instance import app
from server_state import _SERVER_STORE
from utils.chart_helpers import _PLOT_LAYOUT, _AXIS_STYLE


# ── Model Özeti — Markdown editör toolbar + kaydet ────────────────────────

_TOOLBAR_BTN = {
    "backgroundColor": "#1e293b", "color": "#d1d5db", "border": "1px solid #334155",
    "borderRadius": "4px", "padding": "2px 10px", "cursor": "pointer",
    "fontSize": "0.78rem", "marginRight": "4px", "lineHeight": "1.6",
}


def _build_note_editor(existing_note=""):
    """Markdown destekli not editörü (toolbar + textarea + kaydet butonu)."""
    toolbar = html.Div([
        html.Button("B", id="btn-note-bold", n_clicks=0,
                    style={**_TOOLBAR_BTN, "fontWeight": "bold"}),
        html.Button("I", id="btn-note-italic", n_clicks=0,
                    style={**_TOOLBAR_BTN, "fontStyle": "italic"}),
        html.Button("• Liste", id="btn-note-bullet", n_clicks=0,
                    style=_TOOLBAR_BTN),
    ], style={"display": "flex", "alignItems": "center", "marginBottom": "4px"})

    editor = dbc.Textarea(
        id="textarea-model-note",
        value=existing_note,
        placeholder="Model hakkında özet yazabilirsiniz…\n\n"
                    "Toolbar ile **kalın**, *italik*, - madde ekleyebilirsiniz.",
        rows=5,
        style={"backgroundColor": "#111827", "color": "#d1d5db",
               "border": "1px solid #2d3a4f", "fontSize": "0.8rem",
               "fontFamily": "Consolas, 'Courier New', monospace",
               "resize": "vertical"},
    )

    save_btn = dbc.Button("Notu Kaydet", id="btn-save-model-note",
                          color="primary", size="sm",
                          style={"fontSize": "0.75rem", "marginTop": "6px"})

    return html.Div([
        html.Div("Model Özeti", style={
            "color": "#c8cdd8", "fontSize": "0.85rem", "fontWeight": "600",
            "marginBottom": "0.4rem"}),
        toolbar,
        editor,
        html.Div([save_btn, html.Span(id="note-save-status",
                                       style={"marginLeft": "8px", "fontSize": "0.75rem"})],
                 style={"display": "flex", "alignItems": "center"}),
    ], style={"marginBottom": "1rem", "padding": "0.75rem",
              "backgroundColor": "#0d1520", "borderRadius": "6px",
              "border": "1px solid #1e2a3a"})


def _build_note_accordion(note_text):
    """Kaydedilmiş notu accordion olarak göster."""
    if not note_text or not note_text.strip():
        return None
    return dbc.AccordionItem(
        dcc.Markdown(note_text, style={"color": "#d1d5db", "fontSize": "0.82rem",
                                        "lineHeight": "1.7", "padding": "0.5rem"}),
        title="Model Özeti",
        item_id="item-note",
    )


# ── Ortak stiller ──────────────────────────────────────────────────────────────
_TBL_STYLE = dict(
    sort_action="native", page_size=20,
    style_table={"overflowX": "auto"},
    style_header={"backgroundColor": "#161d2e", "color": "#a8b2c2",
                  "fontWeight": "600", "fontSize": "0.7rem",
                  "border": "1px solid #2d3a4f", "textTransform": "uppercase"},
    style_cell={"backgroundColor": "#111827", "color": "#d1d5db",
                "fontSize": "0.78rem", "border": "1px solid #1f2a3c",
                "padding": "5px 8px", "textAlign": "left"},
    style_data_conditional=[
        {"if": {"row_index": "odd"}, "backgroundColor": "#1a2035"}
    ],
)


# ── Yardımcı fonksiyonlar ──────────────────────────────────────────────────────
def _no_model_message():
    return html.Div([
        html.Div("⚠", style={"fontSize": "2.5rem", "color": "#f59e0b",
                               "marginBottom": "0.5rem"}),
        html.P("Önce bir model geliştirilmeli.",
               style={"color": "#c8cdd8", "fontSize": "1.1rem",
                      "fontWeight": "600", "marginBottom": "0.3rem"}),
        html.P("Modelleme sekmesinden model kurun, sonuçlar burada görünecektir.",
               style={"color": "#6b7a99", "fontSize": "0.85rem"}),
    ], style={"textAlign": "center", "padding": "5rem 2rem"})


def _gc(g):
    """Gini değerine göre renk."""
    return "#10b981" if g >= 0.4 else "#f59e0b" if g >= 0.2 else "#ef4444"


def _build_metrics_section(tab_data):
    """Model Metrikleri accordion item'ı."""
    def _mc(v, label, c="#4F8EF7"):
        return dbc.Col(html.Div([
            html.Div(str(v), className="metric-value",
                     style={"color": c, "fontSize": "1.25rem"}),
            html.Div(label, className="metric-label"),
        ], className="metric-card"), width=2)

    def _metric_row(m, title, bg):
        gc = _gc(m["gini"])
        return html.Div([
            html.Div(title, style={"color": "#a8b2c2", "fontSize": "0.72rem",
                                   "fontWeight": "600", "letterSpacing": "0.06em",
                                   "textTransform": "uppercase",
                                   "marginBottom": "0.4rem", "paddingLeft": "0.25rem"}),
            dbc.Row([
                _mc(f"{m['gini']:.4f}", "Gini",      gc),
                _mc(f"{m['auc']:.4f}",  "AUC",       gc),
                _mc(f"{m['ks']:.4f}",   "KS",        "#4F8EF7"),
                _mc(f"{m['f1']:.4f}",   "F1",        "#a78bfa"),
                _mc(f"{m['prec']:.4f}", "Precision", "#a78bfa"),
                _mc(f"{m['rec']:.4f}",  "Recall",    "#a78bfa"),
                _mc(f"{m['n']:,}",      "N",         "#556070"),
            ], className="g-2"),
        ], style={"backgroundColor": bg, "borderRadius": "6px",
                  "padding": "0.6rem 0.5rem", "marginBottom": "0.5rem"})

    metrics = tab_data["metrics"]
    rows = []
    bgs = {"train": "#0d1520", "test": "#0e1624", "oot": "#131c30"}
    labels = {"train": "Train", "test": "Test", "oot": "OOT"}
    for k in ("train", "test", "oot"):
        m = metrics.get(k)
        if m is not None:
            rows.append(_metric_row(m, labels[k], bgs[k]))

    return dbc.AccordionItem(
        html.Div(rows),
        title="Model Metrikleri",
        item_id="item-metrics",
    )


def _build_summary_section(tab_data, algo):
    """Model Özeti accordion item'ı — LR: summary text, Trees: importance table."""
    if algo == "lr" and tab_data.get("lr_summary_text"):
        content = html.Pre(
            tab_data["lr_summary_text"],
            style={
                "backgroundColor": "#0e1117",
                "color": "#c8cdd8",
                "padding": "1rem",
                "borderRadius": "6px",
                "fontSize": "0.72rem",
                "fontFamily": "Consolas, 'Courier New', monospace",
                "overflowX": "auto",
                "border": "1px solid #1f2a3c",
                "whiteSpace": "pre",
                "lineHeight": "1.5",
                "maxHeight": "500px",
                "overflowY": "auto",
            }
        )
    else:
        records = tab_data.get("importance_table", [])
        if not records:
            content = html.Div("Tablo verisi bulunamadı.", style={"color": "#6b7a99"})
        else:
            imp_type = tab_data.get("importance_type", "coef")
            # P-Value renk kodlaması (LR)
            cond = [{"if": {"row_index": "odd"}, "backgroundColor": "#1a2035"}]
            if imp_type == "coef" and "P-Value" in records[0]:
                cond = [
                    {"if": {"filter_query": "{P-Value} < 0.05", "column_id": "P-Value"},
                     "color": "#10b981", "fontWeight": "600"},
                    {"if": {"filter_query": "{P-Value} >= 0.05 && {P-Value} < 0.10",
                            "column_id": "P-Value"},
                     "color": "#f59e0b", "fontWeight": "600"},
                    {"if": {"filter_query": "{P-Value} >= 0.10", "column_id": "P-Value"},
                     "color": "#ef4444"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#1a2035"},
                ]
            content = dash_table.DataTable(
                data=records,
                columns=[{"name": k, "id": k} for k in records[0].keys()],
                style_data_conditional=cond,
                **{k: v for k, v in _TBL_STYLE.items() if k != "style_data_conditional"},
            )

    title = "Model Özeti (Katsayılar)" if algo == "lr" else "Model Özeti (Feature Importance)"
    return dbc.AccordionItem(content, title=title, item_id="item-summary")


def _build_roc_section(tab_data, thr_label):
    """ROC Eğrisi accordion item'ı."""
    accent = tab_data.get("accent", "#4F8EF7")
    roc = tab_data.get("roc_data", {})

    fig = go.Figure()
    traces = []
    if roc.get("train"):
        m = tab_data["metrics"]["train"]
        traces.append((roc["train"], "#556070", f"Train (AUC={m['auc']:.3f})"))
    if roc.get("test"):
        m = tab_data["metrics"]["test"]
        traces.append((roc["test"], accent, f"Test (AUC={m['auc']:.3f})"))
    if roc.get("oot"):
        m = tab_data["metrics"]["oot"]
        traces.append((roc["oot"], "#a78bfa", f"OOT (AUC={m['auc']:.3f})"))

    for rd, col_, nm in traces:
        fig.add_trace(go.Scatter(
            x=rd["fpr"], y=rd["tpr"], mode="lines",
            line=dict(color=col_, width=2), name=nm,
        ))
    fig.add_trace(go.Scatter(
        x=[0, 1], y=[0, 1], mode="lines",
        line=dict(color="#4a5568", dash="dash", width=1), showlegend=False))
    fig.update_layout(
        **_PLOT_LAYOUT,
        title=dict(text=f"ROC Eğrisi — {thr_label}",
                   font=dict(color="#E8EAF0", size=13)),
        xaxis=dict(**_AXIS_STYLE, title="FPR"),
        yaxis=dict(**_AXIS_STYLE, title="TPR"),
        height=380, showlegend=True,
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#8892a4", size=10)),
    )

    return dbc.AccordionItem(
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
        title="ROC Eğrisi",
        item_id="item-roc",
    )


def _build_cm_section(tab_data, thr_label):
    """Confusion Matrix accordion item'ı — Train/Test/OOT yan yana."""
    accent = tab_data.get("accent", "#4F8EF7")
    cms = tab_data.get("confusion_matrices", {})

    cols = []
    for split_name, split_key in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
        cm_data = cms.get(split_key)
        if cm_data is None:
            continue
        tn, fp = cm_data[0][0], cm_data[0][1]
        fn, tp = cm_data[1][0], cm_data[1][1]
        fig = go.Figure(go.Heatmap(
            z=[[tn, fp], [fn, tp]],
            x=["Pred: 0", "Pred: 1"], y=["Actual: 0", "Actual: 1"],
            text=[[str(tn), str(fp)], [str(fn), str(tp)]],
            texttemplate="%{text}",
            textfont=dict(size=14),
            colorscale=[[0, "#0e1117"], [1, accent]], showscale=False,
        ))
        _cm_layout = {k: v for k, v in _PLOT_LAYOUT.items() if k != "margin"}
        fig.update_layout(
            **_cm_layout,
            title=dict(text=f"{split_name}", font=dict(color="#E8EAF0", size=12)),
            height=280, xaxis=dict(side="top"),
            margin=dict(l=60, r=20, t=50, b=20),
        )
        cols.append(dbc.Col(
            dcc.Graph(figure=fig, config={"displayModeBar": False}),
            width=4,
        ))

    if not cols:
        content = html.Div("Confusion matrix verisi bulunamadı.",
                           style={"color": "#6b7a99"})
    else:
        content = html.Div([
            html.Div(thr_label, style={"color": "#7e8fa4", "fontSize": "0.72rem",
                                       "marginBottom": "0.5rem", "fontStyle": "italic"}),
            dbc.Row(cols),
        ])

    return dbc.AccordionItem(content, title="Confusion Matrix", item_id="item-cm")


def _build_shap_section(tab_data):
    """SHAP Beeswarm accordion item'ı."""
    img_b64 = tab_data.get("shap_img_b64")
    if not img_b64:
        return None

    return dbc.AccordionItem(
        html.Img(
            src=f"data:image/png;base64,{img_b64}",
            style={"width": "100%", "maxWidth": "820px",
                   "display": "block", "margin": "0 auto"},
        ),
        title="SHAP Beeswarm",
        item_id="item-shap",
    )


# ── Rating yöntemleri ──────────────────────────────────────────────────────────
_RATING_26_THRESHOLDS = [
    0.00032, 0.00044, 0.0006, 0.00083, 0.00114, 0.00156, 0.00215,
    0.00297, 0.00409, 0.00563, 0.00775, 0.01067, 0.0147, 0.02024,
    0.02788, 0.03839, 0.05287, 0.0728, 0.10026, 0.13807, 0.19014,
    0.26185, 0.36059, 0.49659, 1.0,
]

_RATING_10_THRESHOLDS = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]


def _assign_rating_thresholds(p_arr, thresholds):
    """Proba array'ini eşiklere göre rating'e çevir."""
    import numpy as np
    ratings = np.zeros(len(p_arr), dtype=int)
    for i, p in enumerate(p_arr):
        assigned = len(thresholds) + 1
        for j, thr in enumerate(thresholds):
            if p < thr:
                assigned = j + 1
                break
        ratings[i] = assigned
    return ratings


def _assign_rating_decile(p_arr):
    """Proba array'ini 10 eşit parçaya böl (rank-based decile)."""
    import numpy as np
    import pandas as pd
    # Rank tabanlı: her zaman eşit adetli 10 grup oluşturur
    n = len(p_arr)
    ranks = pd.Series(p_arr).rank(method="first").values
    deciles = np.ceil(ranks / n * 10).astype(int)
    deciles = np.clip(deciles, 1, 10)
    return deciles


def _rating_dist_table_and_chart(p_arr, y_arr, ratings, split_name, accent,
                                 max_rating=25, show_lift=False):
    """Tek bir split için rating dağılımı tablo + grafik oluştur."""
    import numpy as np

    # Rating bazlı count'ları hesapla
    rating_counts = {}
    for r in set(ratings):
        mask = ratings == r
        rating_counts[int(r)] = (int(mask.sum()), int(y_arr[mask].sum()))

    overall_bad_rate = float(y_arr.sum()) / len(y_arr) if len(y_arr) > 0 else 0.0

    # 1'den max_rating'e kadar sabit satırlar
    rows = []
    for r in range(1, max_rating + 1):
        n, bads = rating_counts.get(r, (0, 0))
        goods = n - bads
        bad_rate = round(bads / n * 100, 2) if n > 0 else 0.0
        row = {
            "Rating": r,
            "Count": n,
            "Count%": round(n / len(p_arr) * 100, 2) if len(p_arr) > 0 else 0.0,
            "Bad": bads,
            "Good": goods,
            "Bad Rate %": bad_rate,
        }
        if show_lift:
            decile_br = bads / n if n > 0 else 0.0
            row["Lift"] = round(decile_br / overall_bad_rate, 2) if overall_bad_rate > 0 else 0.0
        rows.append(row)

    tbl = dash_table.DataTable(
        data=rows,
        columns=[{"name": k, "id": k} for k in rows[0].keys()] if rows else [],
        sort_action="native",
        style_table={"overflowX": "auto"},
        style_header={"backgroundColor": "#161d2e", "color": "#a8b2c2",
                      "fontWeight": "600", "fontSize": "0.7rem",
                      "border": "1px solid #2d3a4f", "textTransform": "uppercase"},
        style_cell={"backgroundColor": "#111827", "color": "#d1d5db",
                    "fontSize": "0.78rem", "border": "1px solid #1f2a3c",
                    "padding": "5px 8px", "textAlign": "center"},
        style_data_conditional=[
            {"if": {"row_index": "odd"}, "backgroundColor": "#1a2035"},
            {"if": {"filter_query": "{Bad Rate %} >= 10", "column_id": "Bad Rate %"},
             "color": "#ef4444", "fontWeight": "600"},
            {"if": {"filter_query": "{Bad Rate %} >= 5 && {Bad Rate %} < 10",
                    "column_id": "Bad Rate %"},
             "color": "#f59e0b", "fontWeight": "600"},
        ],
    )

    r_labels = [str(r["Rating"]) for r in rows]
    n_vals = [r["Count"] for r in rows]
    br_vals = [r["Bad Rate %"] for r in rows]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=r_labels, y=n_vals, name="Count",
        marker_color=accent, opacity=0.7,
    ))
    fig.add_trace(go.Scatter(
        x=r_labels, y=br_vals, name="Bad Rate %",
        yaxis="y2", mode="lines+markers",
        line=dict(color="#ef4444", width=2),
        marker=dict(size=5),
    ))
    fig.update_layout(
        **_PLOT_LAYOUT,
        title=dict(text=f"{split_name}",
                   font=dict(color="#E8EAF0", size=12)),
        xaxis=dict(**_AXIS_STYLE, title="Rating", type="category"),
        yaxis=dict(**_AXIS_STYLE, title="Count"),
        yaxis2=dict(title=dict(text="Bad Rate %", font=dict(color="#ef4444")),
                    overlaying="y", side="right",
                    showgrid=False, tickfont=dict(color="#ef4444")),
        height=320, showlegend=True, barmode="overlay",
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#8892a4", size=10)),
    )

    return html.Div([
        html.Div(split_name, style={"color": "#a8b2c2", "fontSize": "0.78rem",
                                    "fontWeight": "600", "textTransform": "uppercase",
                                    "marginBottom": "0.4rem", "marginTop": "0.75rem"}),
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
        tbl,
    ])


def _build_rating_section(tab_data):
    """Rating Dağılımı accordion item'ı — 3 yöntem seçenekli."""
    import numpy as np

    probas = tab_data.get("probabilities", {})
    y_true_data = tab_data.get("y_true", {})

    if not probas or not probas.get("train"):
        return None

    accent = tab_data.get("accent", "#4F8EF7")

    # Her yöntem için split bazlı sonuçları oluştur
    #                   (label,                     thresholds,              max_rating)
    methods = {
        "26-segment": ("26 Segment (PD Eşikleri)", _RATING_26_THRESHOLDS, 25),
        "10-segment": ("10 Segment (Sabit Eşik)", _RATING_10_THRESHOLDS,  10),
        "decile":     ("10 Decile (Eşit Parça)",  None,                   10),
    }

    method_tabs = []
    for method_key, (method_label, thresholds, max_r) in methods.items():
        split_divs = []
        for split_name, split_key in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            p = probas.get(split_key)
            y = y_true_data.get(split_key)
            if p is None or y is None:
                continue

            p_arr = np.array(p)
            y_arr = np.array(y)

            if thresholds is not None:
                ratings = _assign_rating_thresholds(p_arr, thresholds)
            else:
                ratings = _assign_rating_decile(p_arr)

            div = _rating_dist_table_and_chart(p_arr, y_arr, ratings, split_name, accent,
                                               max_rating=max_r,
                                               show_lift=(method_key == "decile"))
            split_divs.append(div)

        if split_divs:
            method_tabs.append(
                dbc.Tab(html.Div(split_divs), label=method_label,
                        tab_id=f"rating-{method_key}", className="tab-content-area")
            )

    if not method_tabs:
        return None

    return dbc.AccordionItem(
        dbc.Tabs(method_tabs, active_tab="rating-26-segment"),
        title="Rating Dağılımı",
        item_id="item-rating",
    )


def _calc_hhi_table(p_arr, thresholds):
    """26 Segment üzerinden rating bazlı HHI tablosu döndürür."""
    import numpy as np
    ratings = _assign_rating_thresholds(p_arr, thresholds)
    n = len(p_arr)
    rows = []
    total_hhi = 0.0
    for r in range(1, 26):
        cnt = int(np.sum(ratings == r))
        share = cnt / n if n > 0 else 0.0
        hhi_contrib = share ** 2
        total_hhi += hhi_contrib
        rows.append({
            "Rating": r,
            "Adet": cnt,
            "Yoğunlaşma": round(share * 100, 2),
            "HHI": round(hhi_contrib, 6),
        })
    rows.append({
        "Rating": "TOPLAM",
        "Adet": n,
        "Yoğunlaşma": 100.0,
        "HHI": round(total_hhi, 6),
    })
    return rows, total_hhi


def _build_hhi_section(tab_data):
    """HHI (Herfindahl-Hirschman Index) accordion — 26 Segment üzerinden."""
    import numpy as np

    probas = tab_data.get("probabilities", {})
    if not probas or not probas.get("train"):
        return None

    split_divs = []
    for split_name, split_key in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
        p = probas.get(split_key)
        if p is None:
            continue
        p_arr = np.array(p)
        if len(p_arr) == 0:
            continue

        rows, total_hhi = _calc_hhi_table(p_arr, _RATING_26_THRESHOLDS)

        if total_hhi < 0.06:
            yorum = "Dengeli"
            yorum_color = "#10b981"
        elif total_hhi < 0.10:
            yorum = "Orta"
            yorum_color = "#f59e0b"
        else:
            yorum = "Yoğun"
            yorum_color = "#ef4444"

        cond = [
            {"if": {"filter_query": '{Rating} = "TOPLAM"'},
             "fontWeight": "700", "backgroundColor": "#161d2e"},
            {"if": {"row_index": "odd"}, "backgroundColor": "#1a2035"},
        ]
        tbl = dash_table.DataTable(
            data=rows,
            columns=[{"name": "Rating", "id": "Rating"},
                     {"name": "Adet", "id": "Adet"},
                     {"name": "Yoğunlaşma %", "id": "Yoğunlaşma"},
                     {"name": "HHI", "id": "HHI"}],
            **{**_TBL_STYLE, "style_data_conditional": cond, "page_size": 30},
        )
        header = html.Div([
            html.Span(split_name, style={"color": "#a8b2c2", "fontWeight": "600",
                                         "fontSize": "0.78rem", "textTransform": "uppercase"}),
            html.Span(f"  HHI = {total_hhi:.6f}  ·  {yorum}",
                      style={"color": yorum_color, "fontSize": "0.75rem", "marginLeft": "0.75rem"}),
        ], style={"marginBottom": "0.4rem", "marginTop": "0.75rem"})
        split_divs.append(html.Div([header, tbl]))

    if not split_divs:
        return None

    note = html.Div(
        f"26 Segment PD eşikleri üzerinden · Mükemmel dağılım ≈ {1/26:.4f}",
        style={"color": "#7e8fa4", "fontSize": "0.72rem", "marginTop": "0.5rem",
               "fontStyle": "italic"},
    )
    split_divs.append(note)
    return dbc.AccordionItem(html.Div(split_divs), title="HHI (Yoğunlaşma)",
                             item_id="item-hhi")


def _build_corr_section(corr_dict):
    """Model değişkenleri arasındaki korelasyon heatmap'i."""
    if not corr_dict:
        return None

    import numpy as np

    vars_ = list(corr_dict.keys())
    n = len(vars_)
    if n < 2:
        return None

    # corr_dict: {var_i: {var_j: val, ...}, ...}
    z_raw = [[corr_dict[r][c] for c in vars_] for r in vars_]
    z_abs = [[abs(v) for v in row] for row in z_raw]

    show_text = n <= 18
    annots = [[f"{z_raw[i][j]:.2f}" for j in range(n)] for i in range(n)] if show_text else None

    fig = go.Figure(data=go.Heatmap(
        z=z_abs, x=vars_, y=vars_,
        text=annots,
        texttemplate="%{text}" if show_text else None,
        textfont=dict(size=9, color="#222"),
        customdata=z_raw,
        colorscale="Reds",
        zmin=0, zmax=1,
        colorbar=dict(
            title=dict(text="|r|", font=dict(color="#555", size=11)),
            thickness=12, len=0.8,
            tickfont=dict(color="#555", size=10),
        ),
        hovertemplate="<b>%{y}</b> × <b>%{x}</b><br>r = %{customdata:.4f}<extra></extra>",
    ))
    cell_px = max(18, min(40, 600 // max(n, 1)))
    fig.update_layout(
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font=dict(family="Inter, Segoe UI, sans-serif", color="#333", size=11),
        margin=dict(l=40, r=20, t=40, b=40),
        title=dict(text=f"Korelasyon Matrisi  ({n} değişken)",
                   font=dict(color="#333", size=13)),
        xaxis=dict(tickangle=-45, tickfont=dict(size=9, color="#555"),
                   showgrid=False, linecolor="#ddd"),
        yaxis=dict(tickfont=dict(size=9, color="#555"),
                   showgrid=False, linecolor="#ddd", autorange="reversed"),
        height=max(320, cell_px * n + 120),
    )

    return dbc.AccordionItem(
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
        title="Korelasyon (Model Değişkenleri)",
        item_id="item-corr",
    )


def _build_woe_dist_section(woe_dist):
    """Model değişkenlerinin WoE bin tabloları + monotonluk bilgisi."""
    if not woe_dist:
        return None

    def _mono_color(m):
        if "Değil" in m or "✗" in m:
            return "#ef4444"
        if "Artan" in m or "Azalan" in m:
            return "#10b981"
        return "#8892a4"

    _iv_style = {"color": "#a78bfa", "fontSize": "0.75rem", "marginRight": "0.5rem"}
    _mono_style_base = {"fontSize": "0.75rem", "fontWeight": "600", "marginRight": "1rem"}

    sections = []
    for var_name, info in woe_dist.items():
        train_rows = info.get("train_table")
        if not train_rows:
            continue

        iv_train = info.get("iv_train", 0)
        mono_train = info.get("monoton", "–")

        # Header: değişken adı + her split için IV ve monotonluk yan yana
        header_children = [
            html.Span(var_name, style={"color": "#E8EAF0", "fontWeight": "700",
                                       "fontSize": "0.85rem", "marginRight": "1rem"}),
            html.Span(f"IV (Train): {iv_train}", style=_iv_style),
            html.Span(mono_train, style={**_mono_style_base, "color": _mono_color(mono_train)}),
        ]

        train_cols = [{"name": k, "id": k} for k in train_rows[0].keys()]
        train_tbl = dash_table.DataTable(data=train_rows, columns=train_cols, **_TBL_STYLE)

        # Test ve OOT tabloları
        test_rows = info.get("test_table")
        iv_test   = info.get("iv_test")
        mono_test = info.get("monoton_test", "–")
        oot_rows  = info.get("oot_table")
        iv_oot    = info.get("iv_oot")
        mono_oot  = info.get("monoton_oot", "–")

        _lbl_style = {"color": "#8892a4", "fontSize": "0.72rem",
                      "fontWeight": "600", "marginBottom": "0.3rem",
                      "textTransform": "uppercase"}

        has_test_tbl = bool(test_rows)
        has_oot_tbl  = bool(oot_rows)
        n_splits = 1 + int(has_test_tbl) + int(has_oot_tbl)

        if n_splits > 1:
            col_w = 12 // n_splits
            cols = [dbc.Col([html.Div("Train", style=_lbl_style), train_tbl], md=col_w)]
            if has_test_tbl:
                _tc = [{"name": k, "id": k} for k in test_rows[0].keys()]
                cols.append(dbc.Col([
                    html.Div("Test", style=_lbl_style),
                    dash_table.DataTable(data=test_rows, columns=_tc, **_TBL_STYLE),
                ], md=col_w))
                header_children.extend([
                    html.Span(f"IV (Test): {iv_test}", style=_iv_style),
                    html.Span(mono_test, style={**_mono_style_base, "color": _mono_color(mono_test)}),
                ])
            if has_oot_tbl:
                _oc = [{"name": k, "id": k} for k in oot_rows[0].keys()]
                cols.append(dbc.Col([
                    html.Div("OOT", style=_lbl_style),
                    dash_table.DataTable(data=oot_rows, columns=_oc, **_TBL_STYLE),
                ], md=col_w))
                header_children.extend([
                    html.Span(f"IV (OOT): {iv_oot}", style=_iv_style),
                    html.Span(mono_oot, style={**_mono_style_base, "color": _mono_color(mono_oot)}),
                ])
            row_content = dbc.Row(cols, className="g-2")
        else:
            row_content = train_tbl

        header = html.Div(header_children, style={"marginBottom": "0.4rem", "marginTop": "0.75rem"})
        sections.append(html.Div([header, row_content]))

    if not sections:
        return None

    return dbc.AccordionItem(
        html.Div(sections),
        title="WoE Dağılımı",
        item_id="item-woe-dist",
    )


def _build_vif_section(vif_data, is_woe=False):
    """VIF (Variance Inflation Factor) tablosu — Train/Test/OOT ayrı."""
    if not vif_data:
        return None

    # Train VIF'e göre sırala
    rows = sorted(vif_data, key=lambda r: (r.get("Train VIF") is None,
                                            -(r.get("Train VIF") or 0)))
    cols = [{"name": k, "id": k} for k in rows[0].keys()]

    vif_cond = list(_TBL_STYLE["style_data_conditional"])
    for cid in [c["id"] for c in cols if "VIF" in c["id"]]:
        vif_cond += [
            {"if": {"filter_query": f"{{{cid}}} >= 10", "column_id": cid},
             "color": "#ef4444", "fontWeight": "600"},
            {"if": {"filter_query": f"{{{cid}}} >= 5 && {{{cid}}} < 10", "column_id": cid},
             "color": "#f59e0b", "fontWeight": "600"},
        ]
    tbl = dash_table.DataTable(
        data=rows, columns=cols,
        **{**_TBL_STYLE, "style_data_conditional": vif_cond},
    )

    children = [tbl]
    if is_woe:
        children.insert(0, html.Div(
            "WoE değişkenleri az sayıda kesikli değer aldığından VIF genellikle düşük çıkar. "
            "Multicollinearity değerlendirmesi için ham değerler sekmesindeki VIF'i kullanın.",
            style={"color": "#7e8fa4", "fontSize": "0.73rem", "fontStyle": "italic",
                   "marginBottom": "0.5rem", "padding": "0.3rem 0.5rem",
                   "backgroundColor": "#0d1520", "borderRadius": "4px",
                   "border": "1px solid #1e2a3a"},
        ))

    return dbc.AccordionItem(children, title="VIF", item_id="item-vif")


def _build_psi_section(psi_data, tab_data=None):
    """PSI accordion — Değişken PSI + Rating PSI."""
    import numpy as np

    has_var_psi = bool(psi_data)
    has_rating_psi = False
    rating_psi_tbl = None

    # ── Rating PSI ─────────────────────────────────────────────────────────────
    if tab_data is not None:
        probas = tab_data.get("probabilities", {})
        p_train = probas.get("train")
        p_oot = probas.get("oot")
        if p_train and p_oot:
            tr_arr = np.array(p_train)
            oot_arr = np.array(p_oot)
            tr_ratings = _assign_rating_thresholds(tr_arr, _RATING_26_THRESHOLDS)
            oot_ratings = _assign_rating_thresholds(oot_arr, _RATING_26_THRESHOLDS)

            eps = 1e-4
            n_tr, n_oot = len(tr_arr), len(oot_arr)
            rows = []
            total_psi = 0.0
            for r in range(1, 26):
                tr_cnt = int((tr_ratings == r).sum())
                oot_cnt = int((oot_ratings == r).sum())
                tr_pct = tr_cnt / n_tr if n_tr > 0 else 0.0
                oot_pct = oot_cnt / n_oot if n_oot > 0 else 0.0
                tr_pct_safe = max(tr_pct, eps)
                oot_pct_safe = max(oot_pct, eps)
                psi_contrib = (oot_pct_safe - tr_pct_safe) * np.log(oot_pct_safe / tr_pct_safe)
                total_psi += psi_contrib
                rows.append({
                    "Rating": r,
                    "Train Adet": tr_cnt,
                    "Train %": round(tr_pct * 100, 2),
                    "OOT Adet": oot_cnt,
                    "OOT %": round(oot_pct * 100, 2),
                    "PSI Katkı": round(float(psi_contrib), 6),
                })
            # Toplam satırı
            rows.append({
                "Rating": "TOPLAM",
                "Train Adet": n_tr,
                "Train %": 100.0,
                "OOT Adet": n_oot,
                "OOT %": 100.0,
                "PSI Katkı": round(float(total_psi), 4),
            })

            r_cols = [{"name": k, "id": k} for k in rows[0].keys()]
            r_cond = list(_TBL_STYLE["style_data_conditional"]) + [
                {"if": {"filter_query": '{PSI Katkı} >= 0.01', "column_id": "PSI Katkı"},
                 "color": "#ef4444", "fontWeight": "600"},
                {"if": {"filter_query": '{Rating} = "TOPLAM"'},
                 "fontWeight": "700", "backgroundColor": "#1a2035",
                 "borderTop": "2px solid #4F8EF7"},
            ]
            rating_psi_tbl = dash_table.DataTable(
                data=rows, columns=r_cols,
                **{**_TBL_STYLE, "style_data_conditional": r_cond,
                   "page_size": 30},
            )
            has_rating_psi = True

    if not has_var_psi and not has_rating_psi:
        return None

    # ── Accordion içeriği ──────────────────────────────────────────────────────
    _section_title = {"color": "#a8b2c2", "fontSize": "0.78rem", "fontWeight": "600",
                      "textTransform": "uppercase", "marginBottom": "0.5rem",
                      "marginTop": "0.75rem"}
    children = []

    if has_var_psi:
        cols = [{"name": k, "id": k} for k in psi_data[0].keys()]
        cond = list(_TBL_STYLE["style_data_conditional"])
        for cid in [c["id"] for c in cols if "PSI" in c["id"]]:
            cond += [
                {"if": {"filter_query": f"{{{cid}}} >= 0.25", "column_id": cid},
                 "color": "#ef4444", "fontWeight": "600"},
                {"if": {"filter_query": f"{{{cid}}} >= 0.10 && {{{cid}}} < 0.25",
                        "column_id": cid},
                 "color": "#f59e0b", "fontWeight": "600"},
            ]
        children.append(html.Div("Değişken PSI", style=_section_title))
        children.append(dash_table.DataTable(
            data=psi_data, columns=cols, **{**_TBL_STYLE, "style_data_conditional": cond},
        ))

    if has_rating_psi:
        children.append(html.Div("Rating PSI (26 Segment)", style={**_section_title, "marginTop": "1.5rem"}))
        children.append(rating_psi_tbl)

    return dbc.AccordionItem(html.Div(children), title="PSI", item_id="item-psi")


def _build_weight_section(tab_data):
    """Ağırlık tablosu — |katsayı| / toplam |katsayılar| * 100."""
    imp_table = tab_data.get("importance_table")
    imp_type  = tab_data.get("importance_type")
    if not imp_table:
        return None

    if imp_type == "coef":
        # const hariç
        var_rows = [r for r in imp_table if r.get("Değişken") != "const"]
        if not var_rows:
            return None
        abs_vals = [abs(r["Katsayı"]) for r in var_rows]
        total = sum(abs_vals) or 1.0
        rows = [{"Değişken": r["Değişken"],
                 "Katsayı": r["Katsayı"],
                 "Ağırlık %": round(abs(r["Katsayı"]) / total * 100, 2)}
                for r in var_rows]
        rows.sort(key=lambda r: r["Ağırlık %"], reverse=True)
        rows.append({"Değişken": "Total", "Katsayı": "–", "Ağırlık %": 100.00})
    elif imp_type == "feature_importance":
        # Zaten yüzdelik var, direkt kullan
        rows = [{"Değişken": r["Değişken"],
                 "Önem (%)": r["Önem (%)"],
                 "Ağırlık %": r["Önem (%)"]}
                for r in imp_table]
        rows.sort(key=lambda r: r["Ağırlık %"], reverse=True)
        rows.append({"Değişken": "Total", "Önem (%)": "–", "Ağırlık %": 100.00})
    else:
        return None

    cols = [{"name": k, "id": k} for k in rows[0].keys()]

    tbl = dash_table.DataTable(data=rows, columns=cols, **_TBL_STYLE)

    return dbc.AccordionItem(tbl, title="Ağırlık", item_id="item-weight")


def _build_describe_section(describe_data):
    """Model değişkenlerinin describe tablosu."""
    if not describe_data:
        return None

    tbl = dash_table.DataTable(
        data=describe_data,
        columns=[{"name": k, "id": k} for k in describe_data[0].keys()],
        **{**_TBL_STYLE, "page_size": 50},
    )

    return dbc.AccordionItem(tbl, title="Describe", item_id="item-describe")


def _build_results_content(tab_data, results, corr_dict=None, woe_dist=None,
                           psi_data=None, vif_data=None,
                           describe_data=None, note_text=None):
    """Tek bir tab (Ham veya WoE) için accordion yapısını oluşturur."""
    algo = results["algo"]
    thr_label = results.get("thr_label", "")

    items = []

    # Model Özeti — en üstte
    note_item = _build_note_accordion(note_text)
    if note_item:
        items.append(note_item)

    items += [
        _build_metrics_section(tab_data),
        _build_summary_section(tab_data, algo),
    ]

    # WoE dağılımı — sadece WoE sekmesi için, Model Özeti'nin hemen altında
    woe_dist_item = _build_woe_dist_section(woe_dist)
    if woe_dist_item:
        items.append(woe_dist_item)

    # Ağırlık — katsayı/önem oranları
    weight_item = _build_weight_section(tab_data)
    if weight_item:
        items.append(weight_item)

    # VIF — Train/Test/OOT
    _is_woe = woe_dist is not None
    vif_item = _build_vif_section(vif_data, is_woe=_is_woe)
    if vif_item:
        items.append(vif_item)

    # PSI — Değişken PSI + Rating PSI
    psi_item = _build_psi_section(psi_data, tab_data)
    if psi_item:
        items.append(psi_item)

    items += [
        _build_roc_section(tab_data, thr_label),
        _build_cm_section(tab_data, thr_label),
    ]

    shap_item = _build_shap_section(tab_data)
    if shap_item:
        items.append(shap_item)

    rating_item = _build_rating_section(tab_data)
    if rating_item:
        items.append(rating_item)

    hhi_item = _build_hhi_section(tab_data)
    if hhi_item:
        items.append(hhi_item)

    corr_item = _build_corr_section(corr_dict)
    if corr_item:
        items.append(corr_item)

    # Describe — model değişkenlerinin profiling tablosu
    desc_item = _build_describe_section(describe_data)
    if desc_item:
        items.append(desc_item)

    return dbc.Accordion(items, always_open=True, active_item=["item-metrics"])


# ── Callback: Sonuç sekmesini render et ────────────────────────────────────────
@app.callback(
    Output("tab-results", "children"),
    Input("main-tabs", "active_tab"),
    Input("store-model-signal", "data"),
    State("store-key", "data"),
    State("input-sql-server", "value"),
    State("input-sql-database", "value"),
    State("dd-sql-driver", "value"),
)
def render_results_tab(active_tab, model_signal, key, _srv, _db, _drv):
    if active_tab != "tab-results":
        return dash.no_update

    if not key:
        return _no_model_message()

    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return _no_model_message()

    results = _SERVER_STORE[cache_key]
    tabs_data = results.get("tabs", {})

    split_info = results.get("split_info", "")
    algo_label = {"lr": "Logistic Regression", "lgbm": "LightGBM",
                  "xgb": "XGBoost", "rf": "Random Forest"}.get(results.get("algo"), "")
    n_vars = len(results.get("model_vars", []))

    header = html.Div([
        html.Span(f"{algo_label}  ·  {n_vars} değişken  ·  {split_info}",
                  style={"color": "#8892a4", "fontSize": "0.78rem"}),
    ], style={"marginBottom": "1rem", "padding": "0.5rem 0.75rem",
              "backgroundColor": "#0d1520", "borderRadius": "6px",
              "border": "1px solid #1e2a3a"})

    woe_corr_data = results.get("corr")  # WoE korelasyon (Train WoE)
    raw_corr_data = results.get("raw_corr")  # Ham korelasyon (Train ham)
    woe_dist_data = results.get("woe_dist")
    woe_psi_data = results.get("psi_data")
    raw_psi_data = results.get("raw_psi_data")
    woe_vif_data = results.get("vif_data")
    raw_vif_data = results.get("raw_vif_data")
    describe_data = results.get("describe_data")
    note_text = results.get("model_note", "")

    sub_tabs = []
    for tab_key, tab_label in [("raw", "Ham Değerler"), ("woe", "WoE Dönüştürülmüş")]:
        tab_data = tabs_data.get(tab_key)
        if tab_data is None:
            continue
        _corr = raw_corr_data if tab_key == "raw" else woe_corr_data
        _psi = raw_psi_data if tab_key == "raw" else woe_psi_data
        _vif = raw_vif_data if tab_key == "raw" else woe_vif_data
        content = _build_results_content(
            tab_data, results,
            corr_dict=_corr,
            woe_dist=woe_dist_data if tab_key == "woe" else None,
            psi_data=_psi,
            vif_data=_vif,
            describe_data=describe_data,
            note_text=note_text,
        )
        sub_tabs.append(dbc.Tab(content, label=tab_label, tab_id=f"res-{tab_key}",
                                className="tab-content-area"))

    if not sub_tabs:
        return _no_model_message()

    # Model özeti editörü — kaydedilmiş not varsa editör boş başlar (accordion'da görünür)
    note_editor = _build_note_editor("")

    # Excel export alanı
    export_bar = html.Div([
        dbc.InputGroup([
            dbc.Input(id="input-export-filename", type="text",
                      placeholder="Dosya adı…", value=f"{algo_label}_{n_vars}vars",
                      style={"maxWidth": "250px", "backgroundColor": "#111827",
                             "color": "#d1d5db", "border": "1px solid #2d3a4f",
                             "fontSize": "0.78rem"}),
            dbc.Button("Excel Kaydet", id="btn-export-excel",
                       color="success", size="sm",
                       style={"fontSize": "0.78rem"}),
        ], size="sm"),
        html.Div(id="excel-save-status", style={"marginTop": "0.4rem"}),
    ], style={"marginBottom": "0.75rem"})

    # SQL & Pickle export bölümü
    model_vars = results.get("model_vars", [])
    all_col_opts = []
    df_orig = _SERVER_STORE.get(key)
    if df_orig is not None:
        import pandas as _pd
        _meta_cols = set()
        _target = results.get("_target", "")
        _date_col = results.get("_date_col", "")
        _seg_col = results.get("_seg_col", "")
        for mc in [_target, _date_col, _seg_col]:
            if mc:
                _meta_cols.add(mc)
        for c in df_orig.columns:
            if c not in _meta_cols and c not in model_vars:
                all_col_opts.append({"label": c, "value": c})

    sql_pickle_section = _build_sql_pickle_section(
        all_col_opts,
        server=_srv or "",
        database=_db or "",
        driver=_drv or "",
    )

    return html.Div([
        header,
        note_editor,
        export_bar,
        dbc.Tabs(sub_tabs, active_tab="res-raw", id="results-sub-tabs"),
        sql_pickle_section,
    ])


# ── Callback: Model Özeti Kaydet ───────────────────────────────────────────
@app.callback(
    Output("note-save-status", "children"),
    Output("textarea-model-note", "value", allow_duplicate=True),
    Input("btn-save-model-note", "n_clicks"),
    State("textarea-model-note", "value"),
    State("store-key", "data"),
    prevent_initial_call=True,
)
def save_model_note(_, note_text, key):
    if not key:
        return "", dash.no_update
    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return "", dash.no_update
    _SERVER_STORE[cache_key]["model_note"] = note_text or ""
    return (
        html.Span("✓ Kaydedildi", style={"color": "#10b981", "fontSize": "0.75rem"}),
        "",  # textarea'yı temizle — not artık accordion'da görünür
    )


# ── Clientside: Toolbar butonları — markdown syntax ekle ─────────────────────
clientside_callback(
    """
    function(n, current) {
        if (!n) return dash_clientside.no_update;
        var el = document.getElementById('textarea-model-note');
        if (!el) return (current || '') + '****';
        var start = el.selectionStart || 0;
        var end = el.selectionEnd || 0;
        var val = current || '';
        var selected = val.substring(start, end);
        if (selected) {
            return val.substring(0, start) + '**' + selected + '**' + val.substring(end);
        }
        return val.substring(0, start) + '**kalın**' + val.substring(end);
    }
    """,
    Output("textarea-model-note", "value", allow_duplicate=True),
    Input("btn-note-bold", "n_clicks"),
    State("textarea-model-note", "value"),
    prevent_initial_call=True,
)

clientside_callback(
    """
    function(n, current) {
        if (!n) return dash_clientside.no_update;
        var el = document.getElementById('textarea-model-note');
        if (!el) return (current || '') + '**';
        var start = el.selectionStart || 0;
        var end = el.selectionEnd || 0;
        var val = current || '';
        var selected = val.substring(start, end);
        if (selected) {
            return val.substring(0, start) + '*' + selected + '*' + val.substring(end);
        }
        return val.substring(0, start) + '*italik*' + val.substring(end);
    }
    """,
    Output("textarea-model-note", "value", allow_duplicate=True),
    Input("btn-note-italic", "n_clicks"),
    State("textarea-model-note", "value"),
    prevent_initial_call=True,
)

clientside_callback(
    """
    function(n, current) {
        if (!n) return dash_clientside.no_update;
        var val = current || '';
        var suffix = val.endsWith('\\n') || val === '' ? '- ' : '\\n- ';
        return val + suffix;
    }
    """,
    Output("textarea-model-note", "value", allow_duplicate=True),
    Input("btn-note-bullet", "n_clicks"),
    State("textarea-model-note", "value"),
    prevent_initial_call=True,
)


# ══════════════════════════════════════════════════════════════════════════════
# SQL'e Bas & Pickle İndir
# ══════════════════════════════════════════════════════════════════════════════

_SQL_INPUT = {"backgroundColor": "#111827", "color": "#d1d5db",
              "border": "1px solid #2d3a4f", "fontSize": "0.78rem"}


def _build_sql_pickle_section(extra_col_opts, server="", database="", driver=""):
    """Sonuç sekmesinin altına SQL push + Pickle export bölümü."""
    return html.Div([
        html.Div([
            html.Span("SQL'e Bas & Pickle", style={
                "fontWeight": "700", "fontSize": "0.95rem", "color": "#f59e0b"}),
            html.Span("  —  Model çıktılarını veritabanına yazın veya profil klasörüne kaydedin",
                       style={"color": "#6b7a99", "fontSize": "0.75rem", "marginLeft": "8px"}),
        ], style={"marginBottom": "0.75rem"}),

        dbc.Row([
            dbc.Col([
                dbc.Label("Server", className="form-label",
                          style={"fontSize": "0.75rem", "color": "#8892a4"}),
                dbc.Input(id="sql-export-server", type="text", size="sm",
                          value=server, style=_SQL_INPUT, readonly=True),
            ], width=4),
            dbc.Col([
                dbc.Label("Database", className="form-label",
                          style={"fontSize": "0.75rem", "color": "#8892a4"}),
                dbc.Input(id="sql-export-database", type="text", size="sm",
                          value=database, style=_SQL_INPUT, readonly=True),
            ], width=4),
            dbc.Col([
                dbc.Label("Driver", className="form-label",
                          style={"fontSize": "0.75rem", "color": "#8892a4"}),
                dbc.Input(id="sql-export-driver", type="text", size="sm",
                          value=driver, style=_SQL_INPUT, readonly=True),
            ], width=4),
        ], className="mb-3"),

        dbc.Row([
            dbc.Col([
                dbc.Label("Ekstra Değişkenler (isteğe bağlı)", className="form-label",
                          style={"fontSize": "0.75rem", "color": "#8892a4"}),
                dcc.Dropdown(id="dd-sql-extra-cols", options=extra_col_opts,
                             multi=True, placeholder="Ekstra kolon seçin…",
                             className="dark-dd", searchable=True),
            ], width=8),
            dbc.Col([
                dbc.Label("Tablo Adı", className="form-label",
                          style={"fontSize": "0.75rem", "color": "#8892a4"}),
                dbc.Input(id="input-sql-table-name", type="text", size="sm",
                          placeholder="tablo_adi", style=_SQL_INPUT),
            ], width=4),
        ], className="mb-3"),

        html.Div([
            dbc.Button("SQL'e Bas", id="btn-sql-push", color="warning", size="sm",
                       style={"fontSize": "0.78rem", "marginRight": "8px"}),
            dbc.Button("Model Pickle Kaydet", id="btn-download-model-pkl", color="info",
                       size="sm", outline=True,
                       style={"fontSize": "0.78rem", "marginRight": "8px"}),
            dbc.Button("OPT Pickle Kaydet", id="btn-download-opt-pkl", color="info",
                       size="sm", outline=True,
                       style={"fontSize": "0.78rem"}),
        ], className="d-flex mb-2"),

        html.Div(id="sql-push-status"),
        html.Div(id="pickle-save-status"),

    ], style={"marginTop": "1.5rem", "padding": "1rem",
              "backgroundColor": "#0f1923", "borderRadius": "8px",
              "border": "2px solid #f59e0b33",
              "boxShadow": "0 0 12px rgba(245,158,11,0.08)"})




# ── Callback: SQL'e Bas ─────────────────────────────────────────────────────
@app.callback(
    Output("sql-push-status", "children"),
    Input("btn-sql-push", "n_clicks"),
    State("store-key", "data"),
    State("sql-export-server", "value"),
    State("sql-export-database", "value"),
    State("sql-export-driver", "value"),
    State("input-sql-table-name", "value"),
    State("dd-sql-extra-cols", "value"),
    State("store-config", "data"),
    State("results-sub-tabs", "active_tab"),
    prevent_initial_call=True,
)
def push_to_sql(_, key, server, database, driver, table_name, extra_cols, config, active_tab):
    import pandas as pd
    import numpy as np
    _A = {"padding": "0.4rem 0.75rem", "fontSize": "0.78rem"}

    if not key:
        return dbc.Alert("Veri yüklenmemiş.", color="warning", style=_A)
    if not server or not database:
        return dbc.Alert("Server ve Database bilgisi gerekli.", color="warning", style=_A)
    if not table_name or not table_name.strip():
        return dbc.Alert("Tablo adı giriniz.", color="warning", style=_A)

    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return dbc.Alert("Önce bir model kurun.", color="warning", style=_A)

    results = _SERVER_STORE[cache_key]
    model_vars = results.get("model_vars", [])
    target_col = results.get("_target", "")
    date_col = results.get("_date_col", "")
    seg_col = results.get("_seg_col", "")
    split_masks = results.get("_split_masks", {})

    df_orig = _SERVER_STORE.get(key)
    if df_orig is None:
        return dbc.Alert("Veri bulunamadı.", color="danger", style=_A)

    from utils.helpers import apply_segment_filter
    seg_val = results.get("_seg_val")
    df_active = apply_segment_filter(df_orig, seg_col, seg_val)

    # WoE DataFrame — merkezi cache'ten
    _pfx = f"{key}_ds_{seg_col}_{seg_val}"
    _train_woe = _SERVER_STORE.get(f"{_pfx}_train_woe")
    _test_woe  = _SERVER_STORE.get(f"{_pfx}_test_woe")
    _oot_woe   = _SERVER_STORE.get(f"{_pfx}_oot_woe")
    woe_parts = [df for df in [_train_woe, _test_woe, _oot_woe] if df is not None]
    woe_df = pd.concat(woe_parts, axis=0).reindex(df_active.index) if woe_parts else None

    # DataFrame oluştur
    out_cols = []
    out_df = pd.DataFrame(index=df_active.index)

    # 1) Ekstra değişkenler (en başta)
    if extra_cols:
        for c in extra_cols:
            if c in df_active.columns:
                out_df[c] = df_active[c].values
                out_cols.append(c)

    # 2) Target, tarih, segment
    for mc, lbl in [(target_col, target_col), (date_col, date_col), (seg_col, seg_col)]:
        if mc and mc in df_active.columns and mc not in out_cols:
            out_df[mc] = df_active[mc].values
            out_cols.append(mc)

    # 3) TRAIN_OOT_FLAG
    train_mask = split_masks.get("train", [])
    test_mask = split_masks.get("test", [])
    oot_mask = split_masks.get("oot", [])
    flag = pd.Series("", index=df_active.index)
    if train_mask and len(train_mask) == len(df_active):
        flag[np.array(train_mask, dtype=bool)] = "TRAIN"
    if test_mask and len(test_mask) == len(df_active):
        flag[np.array(test_mask, dtype=bool)] = "TEST"
    if oot_mask and len(oot_mask) == len(df_active):
        flag[np.array(oot_mask, dtype=bool)] = "OOT"
    out_df["TRAIN_OOT_FLAG"] = flag.values

    # 4) Model değişkenleri — aktif tab'a göre
    _is_woe_tab = active_tab == "res-woe"
    for v in model_vars:
        if v in df_active.columns:
            out_df[v] = df_active[v].values
        if _is_woe_tab and woe_df is not None and v in woe_df.columns:
            out_df[f"{v}_woe"] = woe_df[v].values

    # 5) Predict proba — aktif tab'a göre
    tabs = results.get("tabs", {})
    _proba_pairs = [("woe", "_PROBA_WOE")] if _is_woe_tab else [("raw", "_PROBA_RAW")]
    for tk, suffix in _proba_pairs:
        td = tabs.get(tk)
        if td and "probabilities" in td:
            probs = td["probabilities"]
            prob_arr = np.full(len(df_active), np.nan)
            for split_name, mask_key in [("train", "train"), ("test", "test"), ("oot", "oot")]:
                p = probs.get(split_name)
                m = split_masks.get(mask_key, [])
                if p and m and len(m) == len(df_active):
                    mask_bool = np.array(m, dtype=bool)
                    prob_arr[mask_bool] = p
            out_df[f"PREDICT{suffix}"] = prob_arr

    # SQL push
    try:
        from sqlalchemy import create_engine
        drv = driver or "ODBC Driver 17 for SQL Server"
        conn_str = (f"mssql+pyodbc://{server}/{database}"
                    f"?trusted_connection=yes&driver={drv}&TrustServerCertificate=yes")
        engine = create_engine(conn_str)
        tbl = table_name.strip()
        with engine.connect() as conn:
            out_df.to_sql(name=tbl, con=conn, if_exists="replace", index=True)
        n_rows = len(out_df)
        n_cols = len(out_df.columns)
        return dbc.Alert(
            f"✓ {n_rows:,} satır × {n_cols} kolon → [{database}].[dbo].[{tbl}]",
            color="success", style=_A)
    except Exception as e:
        return dbc.Alert(f"SQL hatası: {e}", color="danger", style=_A)


# ── Callback: Model Pickle → Profil Klasörüne Kaydet ────────────────────────
@app.callback(
    Output("pickle-save-status", "children", allow_duplicate=True),
    Input("btn-download-model-pkl", "n_clicks"),
    State("store-key", "data"),
    State("store-profile-loaded", "data"),
    State("results-sub-tabs", "active_tab"),
    prevent_initial_call=True,
)
def save_model_pickle(_, key, profile_name, active_tab):
    import pickle
    from pathlib import Path
    _A = {"padding": "0.4rem 0.75rem", "fontSize": "0.78rem"}


    if not key:
        return dbc.Alert("Veri yüklenmemiş.", color="warning", style=_A)
    if not profile_name:
        return dbc.Alert("Önce bir profil kaydedin.", color="warning", style=_A)

    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return dbc.Alert("Önce bir model kurun.", color="warning", style=_A)

    results = _SERVER_STORE[cache_key]
    models = results.get("_models", {})
    scalers = results.get("_scalers", {})
    algo = results.get("algo", "model")

    tab_key = "woe" if active_tab == "res-woe" else "raw"
    model_obj = models.get(tab_key)
    scaler_obj = scalers.get(tab_key)

    if model_obj is None:
        return dbc.Alert(f"Seçili sekmede ({tab_key}) model bulunamadı.",
                         color="warning", style=_A)

    bundle = {
        "algo": algo,
        "tab": tab_key,
        "model": model_obj,
        "scaler": scaler_obj,
        "model_vars": results.get("model_vars", []),
        "opt_thr": results.get("opt_thr", 0.5),
    }

    suffix = "_woe" if tab_key == "woe" else "_raw"
    profile_dir = Path(__file__).parent.parent / "profiles" / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    pkl_path = profile_dir / f"{algo}_model{suffix}.pkl"
    with open(pkl_path, "wb") as f:
        pickle.dump(bundle, f, protocol=pickle.HIGHEST_PROTOCOL)

    label = "WoE" if tab_key == "woe" else "Ham"
    return dbc.Alert(f"✓ {label} model pickle kaydedildi → {pkl_path.name}",
                     color="success", style=_A)


# ── Callback: OPT Pickle → Profil Klasörüne Kaydet ──────────────────────────
@app.callback(
    Output("pickle-save-status", "children", allow_duplicate=True),
    Input("btn-download-opt-pkl", "n_clicks"),
    State("store-key", "data"),
    State("store-profile-loaded", "data"),
    State("results-sub-tabs", "active_tab"),
    prevent_initial_call=True,
)
def save_opt_pickle(_, key, profile_name, active_tab):
    import pickle
    from pathlib import Path
    _A = {"padding": "0.4rem 0.75rem", "fontSize": "0.78rem"}

    if not key:
        return dbc.Alert("Veri yüklenmemiş.", color="warning", style=_A)
    if not profile_name:
        return dbc.Alert("Önce bir profil kaydedin.", color="warning", style=_A)

    tab_key = "woe" if active_tab == "res-woe" else "raw"
    if tab_key != "woe":
        return dbc.Alert("OPT pickle sadece WoE sekmesinde kullanılır.",
                         color="info", style=_A)

    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return dbc.Alert("Önce bir model kurun.", color="warning", style=_A)

    results = _SERVER_STORE[cache_key]
    opt_dict = results.get("_opt_dict", {})
    if not opt_dict:
        return dbc.Alert("OPT verisi bulunamadı.", color="warning", style=_A)

    algo = results.get("algo", "model")
    profile_dir = Path(__file__).parent.parent / "profiles" / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    pkl_path = profile_dir / f"{algo}_opt_binning.pkl"
    with open(pkl_path, "wb") as f:
        pickle.dump(opt_dict, f, protocol=pickle.HIGHEST_PROTOCOL)

    return dbc.Alert(f"✓ OPT pickle kaydedildi → {pkl_path.name}",
                     color="success", style=_A)


# ── Excel Export — Ortak Stil Yardımcıları ────────────────────────────────────

def _xl_setup():
    """openpyxl style nesnelerini döndür."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    NAVY       = "0D1B2A"
    HEADER_BG  = "1B2A4A"
    ROW_EVEN   = "F7F9FC"
    ROW_ODD    = "FFFFFF"
    ACCENT     = "3B82F6"
    GREEN      = "059669"
    ORANGE     = "D97706"
    RED        = "DC2626"
    BORDER_CLR = "D1D5DB"
    TITLE_BG   = "1E3A5F"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor=HEADER_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font  = Font(name="Segoe UI", size=10, color="1F2937")
    cell_align = Alignment(horizontal="center", vertical="center")
    cell_align_l = Alignment(horizontal="left", vertical="center")

    title_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=13)
    title_fill = PatternFill("solid", fgColor=TITLE_BG)

    even_fill = PatternFill("solid", fgColor=ROW_EVEN)
    odd_fill  = PatternFill("solid", fgColor=ROW_ODD)

    green_font = Font(name="Segoe UI", bold=True, color=GREEN, size=10)
    orange_font = Font(name="Segoe UI", bold=True, color=ORANGE, size=10)
    red_font   = Font(name="Segoe UI", bold=True, color=RED, size=10)
    accent_font = Font(name="Segoe UI", bold=True, color=ACCENT, size=10)
    bold_font = Font(name="Segoe UI", bold=True, color="1F2937", size=10)

    return dict(
        hdr_font=hdr_font, hdr_fill=hdr_fill, hdr_align=hdr_align,
        cell_font=cell_font, cell_align=cell_align, cell_align_l=cell_align_l,
        border=border, even_fill=even_fill, odd_fill=odd_fill,
        title_font=title_font, title_fill=title_fill,
        green_font=green_font, orange_font=orange_font, red_font=red_font,
        accent_font=accent_font, bold_font=bold_font,
        ACCENT=ACCENT, GREEN=GREEN, ORANGE=ORANGE, RED=RED,
        HEADER_BG=HEADER_BG, ROW_EVEN=ROW_EVEN, NAVY=NAVY,
    )


def _xl_write_title(ws, title, col_count, S):
    """Sheet'in en üstüne koyu renkli başlık bandı yaz."""
    from openpyxl.utils import get_column_letter
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(col_count, 1))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = S["title_font"]
    cell.fill = S["title_fill"]
    cell.alignment = S["hdr_align"]
    ws.row_dimensions[1].height = 30
    # İnce boşluk satırı
    ws.row_dimensions[2].height = 6


def _xl_style_header(ws, start_row, col_count, S):
    """Header satırını stillendir."""
    ws.row_dimensions[start_row].height = 26
    for c in range(1, col_count + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.font = S["hdr_font"]
        cell.fill = S["hdr_fill"]
        cell.alignment = S["hdr_align"]
        cell.border = S["border"]


def _xl_style_data(ws, start_row, end_row, col_count, S,
                   left_align_cols=None, num_fmt_cols=None):
    """Veri satırlarını zebra + border ile stillendir."""
    left_align_cols = left_align_cols or set()
    num_fmt_cols = num_fmt_cols or {}
    for r in range(start_row, end_row + 1):
        fill = S["even_fill"] if (r - start_row) % 2 == 0 else S["odd_fill"]
        ws.row_dimensions[r].height = 22
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = S["cell_font"]
            cell.fill = fill
            cell.alignment = S["cell_align_l"] if c in left_align_cols else S["cell_align"]
            cell.border = S["border"]
            if c in num_fmt_cols:
                cell.number_format = num_fmt_cols[c]


def _xl_auto_width(ws, col_count, min_w=10, max_w=28):
    """Sütun genişliklerini otomatik ayarla."""
    from openpyxl.utils import get_column_letter
    for c in range(1, col_count + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len, max_w)


def _xl_write_df(ws, df, start_row, S, left_align_cols=None, num_fmt_cols=None):
    """DataFrame'i worksheet'e header+data olarak yaz ve stillendir."""
    col_count = len(df.columns)
    # Header
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    _xl_style_header(ws, start_row, col_count, S)
    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    end_row = start_row + len(df)
    if end_row >= start_row + 1:
        _xl_style_data(ws, start_row + 1, end_row, col_count, S,
                       left_align_cols=left_align_cols, num_fmt_cols=num_fmt_cols)
    return end_row


# ── Callback: Excel Export → Profil Klasörüne Kaydet ─────────────────────────
@app.callback(
    Output("excel-save-status", "children"),
    Input("btn-export-excel", "n_clicks"),
    State("input-export-filename", "value"),
    State("results-sub-tabs", "active_tab"),
    State("store-key", "data"),
    State("store-profile-loaded", "data"),
    prevent_initial_call=True,
)
def export_results_excel(_, filename, active_tab, key, profile_name):
    import io
    import numpy as np
    import pandas as pd
    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter
    import plotly.graph_objects as _go

    _A = {"padding": "0.4rem 0.75rem", "fontSize": "0.78rem"}

    if not profile_name:
        return dbc.Alert("Önce bir profil kaydedin.", color="warning", style=_A)

    # ── Plotly → PNG yardımcı ─────────────────────────────────────
    _CHART_BG = "#FAFBFC"
    _CHART_FONT = dict(family="Segoe UI, sans-serif", color="#333", size=11)
    _CHART_MARGIN = dict(l=60, r=30, t=50, b=50)
    _CHART_GRID = dict(showgrid=True, gridcolor="#E5E7EB", zeroline=False)

    def _fig_to_xl_image(fig, width=640, height=400):
        """Plotly fig → openpyxl Image (BytesIO based)."""
        img_bytes = fig.to_image(format="png", width=width, height=height, scale=2)
        img_buf = io.BytesIO(img_bytes)
        return XlImage(img_buf)

    if not key:
        return dbc.Alert("Veri yüklenmemiş.", color="warning", style=_A)

    cache_key = f"{key}_model_results"
    if cache_key not in _SERVER_STORE:
        return dbc.Alert("Önce bir model kurun.", color="warning", style=_A)

    results = _SERVER_STORE[cache_key]
    tabs_data = results.get("tabs", {})
    tab_key = "woe" if active_tab == "res-woe" else "raw"
    tab_data = tabs_data.get(tab_key)
    if tab_data is None:
        return dbc.Alert("Sekme verisi bulunamadı.", color="warning", style=_A)

    fname = (filename or "sonuc").strip()
    if not fname:
        fname = "sonuc"
    # WoE seçiliyse dosya adına _woe ekle
    if tab_key == "woe" and not fname.lower().endswith("_woe"):
        fname += "_woe"

    algo = results.get("algo", "")
    algo_label = {"lr": "Logistic Regression", "lgbm": "LightGBM",
                  "xgb": "XGBoost", "rf": "Random Forest"}.get(algo, algo)
    model_vars = results.get("model_vars", [])
    split_info = results.get("split_info", "")
    thr_label = results.get("thr_label", "")

    S = _xl_setup()
    wb = Workbook()
    # Varsayılan boş sheet'i kaldır
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Özet (Dashboard sayfası)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Özet")
    ws.sheet_properties.tabColor = S["NAVY"]

    # Başlık bandı
    ws.merge_cells("A1:H1")
    c1 = ws.cell(row=1, column=1, value=f"Model Raporu — {algo_label}")
    c1.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=16)
    c1.fill = PatternFill("solid", fgColor=S["NAVY"])
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Alt bilgi
    info_items = [
        ("Model:", algo_label),
        ("Değişken Sayısı:", str(len(model_vars))),
        ("Bölünme:", split_info),
        ("Eşik:", thr_label),
        ("Sekme:", "WoE Dönüştürülmüş" if tab_key == "woe" else "Ham Değerler"),
    ]
    for i, (lbl, val) in enumerate(info_items):
        r = 3 + i
        ws.cell(row=r, column=1, value=lbl).font = S["bold_font"]
        ws.cell(row=r, column=2, value=val).font = S["cell_font"]
        ws.cell(row=r, column=1).alignment = S["cell_align_l"]
        ws.cell(row=r, column=2).alignment = S["cell_align_l"]

    # Model Özeti
    note_text = results.get("model_note", "")
    r_start = 3 + len(info_items) + 1
    if note_text and note_text.strip():
        ws.cell(row=r_start, column=1, value="Model Özeti:").font = S["bold_font"]
        ws.cell(row=r_start, column=1).alignment = S["cell_align_l"]
        r_start += 1
        # Markdown → düz metin (basit temizlik)
        clean = note_text.replace("**", "").replace("*", "")
        for nl, nline in enumerate(clean.split("\n")):
            ws.cell(row=r_start + nl, column=1, value=nline).font = S["cell_font"]
            ws.cell(row=r_start + nl, column=1).alignment = S["cell_align_l"]
        r_start += len(clean.split("\n")) + 1

    # Değişken listesi
    ws.cell(row=r_start, column=1, value="Model Değişkenleri:").font = S["bold_font"]
    ws.cell(row=r_start, column=1).alignment = S["cell_align_l"]
    for vi, vn in enumerate(model_vars):
        ws.cell(row=r_start + 1 + vi, column=1, value=f"  {vi+1}. {vn}").font = S["cell_font"]
        ws.cell(row=r_start + 1 + vi, column=1).alignment = S["cell_align_l"]

    # Metrikler tablosu — Özet sayfasında da göster
    metrics = tab_data.get("metrics", {})
    if metrics:
        mr = r_start + len(model_vars) + 3
        ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=8)
        ws.cell(row=mr, column=1, value="Model Performansı").font = Font(
            name="Segoe UI", bold=True, color=S["ACCENT"], size=12)
        ws.cell(row=mr, column=1).alignment = S["cell_align_l"]
        mr += 1

        m_headers = ["Split", "AUC", "Gini", "KS", "F1", "Precision", "Recall", "N"]
        for ci, h in enumerate(m_headers, 1):
            ws.cell(row=mr, column=ci, value=h)
        _xl_style_header(ws, mr, len(m_headers), S)

        for si, (sn, sk) in enumerate([("Train", "train"), ("Test", "test"), ("OOT", "oot")]):
            m = metrics.get(sk)
            if m is None:
                continue
            dr = mr + 1 + si
            vals = [sn, m.get("auc"), m.get("gini"), m.get("ks"),
                    m.get("f1"), m.get("prec"), m.get("rec"), m.get("n")]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=dr, column=ci, value=v)
                cell.border = S["border"]
                cell.alignment = S["cell_align"]
                cell.font = S["cell_font"]
                fill = S["even_fill"] if si % 2 == 0 else S["odd_fill"]
                cell.fill = fill
                if ci >= 2 and ci <= 7 and isinstance(v, (int, float)):
                    cell.number_format = "0.0000"
                    # Gini renklendirme
                    if ci == 3:
                        if v >= 0.4:
                            cell.font = S["green_font"]
                        elif v >= 0.2:
                            cell.font = S["orange_font"]
                        else:
                            cell.font = S["red_font"]
                if ci == 8 and isinstance(v, (int, float)):
                    cell.number_format = "#,##0"

    _xl_auto_width(ws, 8, min_w=12)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Metrikler (detaylı)
    # ═══════════════════════════════════════════════════════════════════════════
    if metrics:
        ws_m = wb.create_sheet("Metrikler")
        ws_m.sheet_properties.tabColor = S["ACCENT"]
        _xl_write_title(ws_m, "Model Metrikleri", 8, S)

        m_rows = []
        for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            m = metrics.get(sk)
            if m is None:
                continue
            m_rows.append({
                "Split": sn, "AUC": m.get("auc"), "Gini": m.get("gini"),
                "KS": m.get("ks"), "F1": m.get("f1"),
                "Precision": m.get("prec"), "Recall": m.get("rec"),
                "N": m.get("n"),
            })
        df_m = pd.DataFrame(m_rows)
        end_r = _xl_write_df(ws_m, df_m, 3, S, left_align_cols={1},
                             num_fmt_cols={2: "0.0000", 3: "0.0000", 4: "0.0000",
                                           5: "0.0000", 6: "0.0000", 7: "0.0000",
                                           8: "#,##0"})

        # Gini renklendirme
        for r in range(4, end_r + 1):
            gini_cell = ws_m.cell(row=r, column=3)
            if isinstance(gini_cell.value, (int, float)):
                if gini_cell.value >= 0.4:
                    gini_cell.font = S["green_font"]
                elif gini_cell.value >= 0.2:
                    gini_cell.font = S["orange_font"]
                else:
                    gini_cell.font = S["red_font"]

        # Metrik karşılaştırma bar chart (Plotly PNG)
        if len(m_rows) >= 2:
            chart_metrics = ["AUC", "Gini", "KS", "F1"]
            colors = ["#5B7FA5", f"#{S['ACCENT']}", "#8B5CF6"]
            fig = _go.Figure()
            for mi, mr in enumerate(m_rows):
                fig.add_trace(_go.Bar(
                    name=mr["Split"], x=chart_metrics,
                    y=[mr[cm] for cm in chart_metrics],
                    marker_color=colors[mi % len(colors)],
                ))
            fig.update_layout(
                barmode="group", title="Metrik Karşılaştırma",
                paper_bgcolor=_CHART_BG, plot_bgcolor=_CHART_BG,
                font=_CHART_FONT, margin=_CHART_MARGIN,
                yaxis=dict(range=[0, 1], **_CHART_GRID),
                legend=dict(orientation="h", y=-0.15),
            )
            img = _fig_to_xl_image(fig, 640, 380)
            ws_m.add_image(img, f"A{end_r + 3}")

        _xl_auto_width(ws_m, 8)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Katsayılar / Feature Importance
    # ═══════════════════════════════════════════════════════════════════════════
    imp_table = tab_data.get("importance_table")
    imp_type = tab_data.get("importance_type")
    if imp_table:
        ws_k = wb.create_sheet("Katsayılar")
        ws_k.sheet_properties.tabColor = "6366F1"
        title_txt = "Katsayılar" if imp_type == "coef" else "Feature Importance"
        _xl_write_title(ws_k, title_txt, len(imp_table[0]), S)
        df_imp = pd.DataFrame(imp_table)
        end_r = _xl_write_df(ws_k, df_imp, 3, S, left_align_cols={1})

        # P-Value renklendirme (LR)
        if imp_type == "coef" and "P-Value" in df_imp.columns:
            pv_col = list(df_imp.columns).index("P-Value") + 1
            for r in range(4, end_r + 1):
                cell = ws_k.cell(row=r, column=pv_col)
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0.05:
                        cell.font = S["green_font"]
                    elif cell.value < 0.10:
                        cell.font = S["orange_font"]
                    else:
                        cell.font = S["red_font"]

        # Feature importance bar chart (Plotly PNG)
        if imp_type == "feature_importance" and len(imp_table) > 1:
            imp_col_name = "Önem (%)" if "Önem (%)" in df_imp.columns else list(df_imp.columns)[-1]
            sorted_imp = df_imp.sort_values(imp_col_name, ascending=True)
            fig = _go.Figure(_go.Bar(
                x=sorted_imp[imp_col_name].tolist(),
                y=sorted_imp["Değişken"].tolist(),
                orientation="h", marker_color=f"#{S['ACCENT']}",
            ))
            fig.update_layout(
                title="Feature Importance",
                paper_bgcolor=_CHART_BG, plot_bgcolor=_CHART_BG,
                font=_CHART_FONT, margin=dict(l=140, r=30, t=50, b=40),
                xaxis=dict(title=imp_col_name, **_CHART_GRID),
                yaxis=dict(**_CHART_GRID),
            )
            h = max(350, len(imp_table) * 28)
            img = _fig_to_xl_image(fig, 660, h)
            ws_k.add_image(img, f"A{end_r + 3}")

        _xl_auto_width(ws_k, len(df_imp.columns))

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Model Özeti (LR summary text)
    # ═══════════════════════════════════════════════════════════════════════════
    lr_text = tab_data.get("lr_summary_text")
    if lr_text:
        ws_lr = wb.create_sheet("Model Özeti")
        ws_lr.sheet_properties.tabColor = "8B5CF6"
        _xl_write_title(ws_lr, "Logistic Regression — Model Özeti", 1, S)
        mono_font = Font(name="Consolas", size=9, color="1F2937")
        for li, line in enumerate(lr_text.split("\n"), 3):
            cell = ws_lr.cell(row=li, column=1, value=line)
            cell.font = mono_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_lr.column_dimensions["A"].width = 100

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Ağırlık
    # ═══════════════════════════════════════════════════════════════════════════
    if imp_table:
        w_rows = []
        if imp_type == "coef":
            var_rows = [r for r in imp_table if r.get("Değişken") != "const"]
            if var_rows:
                abs_vals = [abs(r["Katsayı"]) for r in var_rows]
                total = sum(abs_vals) or 1.0
                w_rows = [{"Değişken": r["Değişken"], "Katsayı": r["Katsayı"],
                           "Ağırlık %": round(abs(r["Katsayı"]) / total * 100, 2)}
                          for r in var_rows]
                w_rows.sort(key=lambda r: r["Ağırlık %"], reverse=True)
                w_rows.append({"Değişken": "Total", "Katsayı": "", "Ağırlık %": 100.0})
        elif imp_type == "feature_importance":
            w_rows = [{"Değişken": r["Değişken"],
                       "Önem (%)": r.get("Önem (%)", 0),
                       "Ağırlık %": r.get("Önem (%)", 0)}
                      for r in imp_table]
            w_rows.sort(key=lambda r: r["Ağırlık %"], reverse=True)
            w_rows.append({"Değişken": "Total", "Önem (%)": "", "Ağırlık %": 100.0})

        if w_rows:
            ws_w = wb.create_sheet("Ağırlık")
            ws_w.sheet_properties.tabColor = "F59E0B"
            _xl_write_title(ws_w, "Değişken Ağırlıkları", len(w_rows[0]), S)
            df_w = pd.DataFrame(w_rows)
            end_r = _xl_write_df(ws_w, df_w, 3, S, left_align_cols={1},
                                 num_fmt_cols={len(df_w.columns): "0.00"})

            # Total satırını kalın yap
            total_row = end_r
            for c in range(1, len(df_w.columns) + 1):
                cell = ws_w.cell(row=total_row, column=c)
                cell.font = S["bold_font"]
                cell.fill = PatternFill("solid", fgColor="E5E7EB")

            # Ağırlık pie chart (Plotly PNG)
            if len(w_rows) > 2:
                labels = [r["Değişken"] for r in w_rows[:-1]]
                values = [r["Ağırlık %"] for r in w_rows[:-1]]
                fig = _go.Figure(_go.Pie(
                    labels=labels, values=values,
                    textinfo="label+percent", textposition="outside",
                    hole=0.35, marker=dict(line=dict(color="#fff", width=1)),
                ))
                fig.update_layout(
                    title="Değişken Ağırlık Dağılımı",
                    paper_bgcolor=_CHART_BG, font=_CHART_FONT,
                    margin=dict(l=20, r=20, t=50, b=20),
                    showlegend=False,
                )
                img = _fig_to_xl_image(fig, 520, 420)
                ws_w.add_image(img, f"D3")

            _xl_auto_width(ws_w, len(df_w.columns))

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 6 — VIF
    # ═══════════════════════════════════════════════════════════════════════════
    vif_data = results.get("vif_data")
    if vif_data:
        ws_v = wb.create_sheet("VIF")
        ws_v.sheet_properties.tabColor = "EF4444"
        _xl_write_title(ws_v, "VIF — Variance Inflation Factor (WoE)", len(vif_data[0]), S)

        df_vif = pd.DataFrame(sorted(vif_data, key=lambda r: (r.get("Train VIF") is None,
                                                                -(r.get("Train VIF") or 0))))
        _vif_num_cols = {i + 1: "0.00" for i, cn in enumerate(df_vif.columns) if "VIF" in cn}
        end_r = _xl_write_df(ws_v, df_vif, 3, S, left_align_cols={1},
                             num_fmt_cols=_vif_num_cols)

        # VIF renklendirme
        _vif_col_idxs = [i + 1 for i, cn in enumerate(df_vif.columns) if "VIF" in cn]
        for r in range(4, end_r + 1):
            for vc in _vif_col_idxs:
                cell = ws_v.cell(row=r, column=vc)
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 10:
                        cell.font = S["red_font"]
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    elif cell.value >= 5:
                        cell.font = S["orange_font"]
                        cell.fill = PatternFill("solid", fgColor="FEF3C7")

        # VIF eşik açıklama
        legend_r = end_r + 2
        ws_v.cell(row=legend_r, column=1, value="VIF ≥ 10").font = S["red_font"]
        ws_v.cell(row=legend_r, column=2, value="Yüksek çoklu bağlantı").font = S["cell_font"]
        ws_v.cell(row=legend_r + 1, column=1, value="VIF ≥ 5").font = S["orange_font"]
        ws_v.cell(row=legend_r + 1, column=2, value="Orta çoklu bağlantı").font = S["cell_font"]
        ws_v.cell(row=legend_r + 2, column=1, value="VIF < 5").font = S["green_font"]
        ws_v.cell(row=legend_r + 2, column=2, value="Sorun yok").font = S["cell_font"]

        _xl_auto_width(ws_v, 2, min_w=14)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 7 — PSI
    # ═══════════════════════════════════════════════════════════════════════════
    psi_data = results.get("psi_data")
    if psi_data:
        ws_p = wb.create_sheet("PSI")
        ws_p.sheet_properties.tabColor = "D97706"
        _xl_write_title(ws_p, "PSI — Population Stability Index", len(psi_data[0]), S)

        df_psi = pd.DataFrame(psi_data)
        end_r = _xl_write_df(ws_p, df_psi, 3, S, left_align_cols={1},
                             num_fmt_cols={c: "0.0000" for c in range(2, len(df_psi.columns) + 1)})

        # PSI renklendirme
        psi_cols = [i + 1 for i, cn in enumerate(df_psi.columns) if "PSI" in cn]
        for r in range(4, end_r + 1):
            for pc in psi_cols:
                cell = ws_p.cell(row=r, column=pc)
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0.25:
                        cell.font = S["red_font"]
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    elif cell.value >= 0.10:
                        cell.font = S["orange_font"]
                        cell.fill = PatternFill("solid", fgColor="FEF3C7")

        legend_r = end_r + 2
        ws_p.cell(row=legend_r, column=1, value="PSI ≥ 0.25").font = S["red_font"]
        ws_p.cell(row=legend_r, column=2, value="Anlamlı kayma").font = S["cell_font"]
        ws_p.cell(row=legend_r + 1, column=1, value="PSI ≥ 0.10").font = S["orange_font"]
        ws_p.cell(row=legend_r + 1, column=2, value="Hafif kayma").font = S["cell_font"]
        ws_p.cell(row=legend_r + 2, column=1, value="PSI < 0.10").font = S["green_font"]
        ws_p.cell(row=legend_r + 2, column=2, value="Stabil").font = S["cell_font"]

        _xl_auto_width(ws_p, len(df_psi.columns))

        # ── Rating PSI (26 Segment) — aynı sheet'e devam ──────────────────────
        probas = tab_data.get("probabilities", {})
        p_train = probas.get("train")
        p_oot = probas.get("oot")
        if p_train and p_oot:
            import numpy as np
            tr_arr = np.array(p_train)
            oot_arr = np.array(p_oot)
            tr_ratings = _assign_rating_thresholds(tr_arr, _RATING_26_THRESHOLDS)
            oot_ratings = _assign_rating_thresholds(oot_arr, _RATING_26_THRESHOLDS)
            eps = 1e-4
            n_tr, n_oot = len(tr_arr), len(oot_arr)
            rpsi_rows = []
            total_psi = 0.0
            for r in range(1, 26):
                tr_cnt = int((tr_ratings == r).sum())
                oot_cnt = int((oot_ratings == r).sum())
                tr_pct = tr_cnt / n_tr if n_tr > 0 else 0.0
                oot_pct = oot_cnt / n_oot if n_oot > 0 else 0.0
                tr_safe = max(tr_pct, eps)
                oot_safe = max(oot_pct, eps)
                psi_c = float((oot_safe - tr_safe) * np.log(oot_safe / tr_safe))
                total_psi += psi_c
                rpsi_rows.append({
                    "Rating": r, "Train Adet": tr_cnt,
                    "Train %": round(tr_pct * 100, 2),
                    "OOT Adet": oot_cnt, "OOT %": round(oot_pct * 100, 2),
                    "PSI Katkı": round(psi_c, 6),
                })
            rpsi_rows.append({
                "Rating": "TOPLAM", "Train Adet": n_tr,
                "Train %": 100.0, "OOT Adet": n_oot,
                "OOT %": 100.0, "PSI Katkı": round(total_psi, 4),
            })

            rpsi_start = legend_r + 4
            ws_p.cell(row=rpsi_start, column=1, value="Rating PSI (26 Segment)").font = Font(
                name="Segoe UI", bold=True, color="FFFFFF", size=12)
            ws_p.cell(row=rpsi_start, column=1).fill = PatternFill("solid", fgColor="1E293B")
            for ci in range(2, 7):
                ws_p.cell(row=rpsi_start, column=ci).fill = PatternFill("solid", fgColor="1E293B")

            df_rpsi = pd.DataFrame(rpsi_rows)
            end_rpsi = _xl_write_df(ws_p, df_rpsi, rpsi_start + 1, S, left_align_cols={1},
                                    num_fmt_cols={6: "0.000000"})

            # TOPLAM satırını kalın yap
            for ci in range(1, 7):
                cell = ws_p.cell(row=end_rpsi, column=ci)
                cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor="1a2035")

            # PSI Katkı renklendirme
            for r in range(rpsi_start + 2, end_rpsi):
                cell = ws_p.cell(row=r, column=6)
                if isinstance(cell.value, (int, float)) and cell.value >= 0.01:
                    cell.font = S["red_font"]

            _xl_auto_width(ws_p, 6, min_w=12)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 8 — ROC Eğrisi (veri + grafik)
    # ═══════════════════════════════════════════════════════════════════════════
    roc_data = tab_data.get("roc_data", {})
    has_roc = any(roc_data.get(k) for k in ("train", "test", "oot"))
    if has_roc:
        ws_roc = wb.create_sheet("ROC")
        ws_roc.sheet_properties.tabColor = "10B981"
        _xl_write_title(ws_roc, f"ROC Eğrisi — {thr_label}", 6, S)

        # Her split için FPR/TPR sütunları yan yana
        col_offset = 1
        split_cols = {}  # {split: (fpr_col, tpr_col)}
        for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            rd = roc_data.get(sk)
            if rd is None:
                continue
            fpr_col = col_offset
            tpr_col = col_offset + 1
            split_cols[sn] = (fpr_col, tpr_col)
            # Header
            ws_roc.cell(row=3, column=fpr_col, value=f"FPR ({sn})")
            ws_roc.cell(row=3, column=tpr_col, value=f"TPR ({sn})")
            _xl_style_header(ws_roc, 3, tpr_col, S)
            # Data
            fpr_list = rd["fpr"]
            tpr_list = rd["tpr"]
            # Her 5. noktayı al (ROC çok fazla nokta olabilir)
            step = max(1, len(fpr_list) // 100)
            sampled = list(range(0, len(fpr_list), step))
            if sampled[-1] != len(fpr_list) - 1:
                sampled.append(len(fpr_list) - 1)

            for ri, idx in enumerate(sampled):
                ws_roc.cell(row=4 + ri, column=fpr_col, value=round(fpr_list[idx], 6))
                ws_roc.cell(row=4 + ri, column=tpr_col, value=round(tpr_list[idx], 6))
                fill = S["even_fill"] if ri % 2 == 0 else S["odd_fill"]
                for cc in (fpr_col, tpr_col):
                    c_ = ws_roc.cell(row=4 + ri, column=cc)
                    c_.font = S["cell_font"]
                    c_.fill = fill
                    c_.border = S["border"]
                    c_.alignment = S["cell_align"]
                    c_.number_format = "0.000000"

            col_offset += 3  # 2 sütun + 1 boşluk

        # ROC chart (Plotly PNG)
        _sn_to_sk = {"Train": "train", "Test": "test", "OOT": "oot"}
        max_data_rows = 0
        for sn in split_cols:
            rd = roc_data.get(_sn_to_sk[sn])
            step = max(1, len(rd["fpr"]) // 100)
            sampled = list(range(0, len(rd["fpr"]), step))
            if sampled[-1] != len(rd["fpr"]) - 1:
                sampled.append(len(rd["fpr"]) - 1)
            max_data_rows = max(max_data_rows, len(sampled))

        line_colors = {"Train": "#5B7FA5", "Test": f"#{S['ACCENT']}", "OOT": "#8B5CF6"}
        fig = _go.Figure()
        for sn in split_cols:
            rd = roc_data.get(_sn_to_sk[sn])
            m = metrics.get(_sn_to_sk[sn], {})
            auc_val = m.get("auc", 0)
            fig.add_trace(_go.Scatter(
                x=rd["fpr"], y=rd["tpr"], mode="lines",
                name=f"{sn} (AUC={auc_val:.4f})",
                line=dict(color=line_colors.get(sn, "#3B82F6"), width=2),
            ))
        fig.add_trace(_go.Scatter(
            x=[0, 1], y=[0, 1], mode="lines",
            line=dict(color="#999", dash="dash", width=1), showlegend=False,
        ))
        fig.update_layout(
            title=f"ROC Eğrisi — {thr_label}",
            paper_bgcolor=_CHART_BG, plot_bgcolor=_CHART_BG,
            font=_CHART_FONT, margin=_CHART_MARGIN,
            xaxis=dict(title="FPR", range=[0, 1], **_CHART_GRID),
            yaxis=dict(title="TPR", range=[0, 1], **_CHART_GRID),
            legend=dict(x=0.55, y=0.05, bgcolor="rgba(255,255,255,0.8)"),
        )
        img = _fig_to_xl_image(fig, 640, 500)
        ws_roc.add_image(img, f"A{4 + max_data_rows + 2}")

        _xl_auto_width(ws_roc, col_offset, min_w=14)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 9 — Confusion Matrix (görsel heatmap benzeri)
    # ═══════════════════════════════════════════════════════════════════════════
    cm_all = tab_data.get("confusion_matrices", {})
    has_cm = any(cm_all.get(k) for k in ("train", "test", "oot"))
    if has_cm:
        ws_cm = wb.create_sheet("Confusion Matrix")
        ws_cm.sheet_properties.tabColor = "EC4899"
        _xl_write_title(ws_cm, f"Confusion Matrix — {thr_label}", 12, S)

        cm_colors = {
            "TP": ("DCFCE7", "166534"),  # yeşil bg, koyu yeşil text
            "TN": ("DBEAFE", "1E40AF"),  # mavi bg, koyu mavi text
            "FP": ("FEF3C7", "92400E"),  # sarı bg, koyu kahve text
            "FN": ("FEE2E2", "991B1B"),  # kırmızı bg, koyu kırmızı text
        }

        col_offset = 1
        for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            cm = cm_all.get(sk)
            if cm is None:
                continue
            tn, fp = cm[0][0], cm[0][1]
            fn, tp = cm[1][0], cm[1][1]
            total = tn + fp + fn + tp
            acc = (tn + tp) / total * 100 if total > 0 else 0

            # Başlık
            r = 3
            ws_cm.merge_cells(start_row=r, start_column=col_offset,
                              end_row=r, end_column=col_offset + 2)
            title_cell = ws_cm.cell(row=r, column=col_offset, value=f"{sn}  (Acc: {acc:.1f}%)")
            title_cell.font = Font(name="Segoe UI", bold=True, color=S["ACCENT"], size=11)
            title_cell.alignment = Alignment(horizontal="center")

            # Sütun başlıkları
            r = 4
            ws_cm.cell(row=r, column=col_offset, value="").border = S["border"]
            ws_cm.cell(row=r, column=col_offset + 1, value="Pred: 0")
            ws_cm.cell(row=r, column=col_offset + 2, value="Pred: 1")
            for cc in range(col_offset, col_offset + 3):
                c_ = ws_cm.cell(row=4, column=cc)
                c_.font = S["hdr_font"]
                c_.fill = S["hdr_fill"]
                c_.alignment = S["hdr_align"]
                c_.border = S["border"]

            # Actual: 0 satırı
            r = 5
            ws_cm.cell(row=r, column=col_offset, value="Actual: 0").font = S["bold_font"]
            ws_cm.cell(row=r, column=col_offset).fill = S["hdr_fill"]
            ws_cm.cell(row=r, column=col_offset).font = S["hdr_font"]
            ws_cm.cell(row=r, column=col_offset).border = S["border"]
            ws_cm.cell(row=r, column=col_offset).alignment = S["hdr_align"]

            tn_cell = ws_cm.cell(row=r, column=col_offset + 1, value=tn)
            tn_cell.fill = PatternFill("solid", fgColor=cm_colors["TN"][0])
            tn_cell.font = Font(name="Segoe UI", bold=True, color=cm_colors["TN"][1], size=13)
            tn_cell.alignment = Alignment(horizontal="center", vertical="center")
            tn_cell.border = S["border"]
            tn_cell.number_format = "#,##0"

            fp_cell = ws_cm.cell(row=r, column=col_offset + 2, value=fp)
            fp_cell.fill = PatternFill("solid", fgColor=cm_colors["FP"][0])
            fp_cell.font = Font(name="Segoe UI", bold=True, color=cm_colors["FP"][1], size=13)
            fp_cell.alignment = Alignment(horizontal="center", vertical="center")
            fp_cell.border = S["border"]
            fp_cell.number_format = "#,##0"

            # Actual: 1 satırı
            r = 6
            ws_cm.cell(row=r, column=col_offset, value="Actual: 1").font = S["bold_font"]
            ws_cm.cell(row=r, column=col_offset).fill = S["hdr_fill"]
            ws_cm.cell(row=r, column=col_offset).font = S["hdr_font"]
            ws_cm.cell(row=r, column=col_offset).border = S["border"]
            ws_cm.cell(row=r, column=col_offset).alignment = S["hdr_align"]

            fn_cell = ws_cm.cell(row=r, column=col_offset + 1, value=fn)
            fn_cell.fill = PatternFill("solid", fgColor=cm_colors["FN"][0])
            fn_cell.font = Font(name="Segoe UI", bold=True, color=cm_colors["FN"][1], size=13)
            fn_cell.alignment = Alignment(horizontal="center", vertical="center")
            fn_cell.border = S["border"]
            fn_cell.number_format = "#,##0"

            tp_cell = ws_cm.cell(row=r, column=col_offset + 2, value=tp)
            tp_cell.fill = PatternFill("solid", fgColor=cm_colors["TP"][0])
            tp_cell.font = Font(name="Segoe UI", bold=True, color=cm_colors["TP"][1], size=13)
            tp_cell.alignment = Alignment(horizontal="center", vertical="center")
            tp_cell.border = S["border"]
            tp_cell.number_format = "#,##0"

            # Sütun genişlikleri
            for cc in range(col_offset, col_offset + 3):
                ws_cm.column_dimensions[get_column_letter(cc)].width = 16
            ws_cm.row_dimensions[5].height = 36
            ws_cm.row_dimensions[6].height = 36

            col_offset += 4  # 3 sütun + 1 boşluk

        # Detaylı tablo altına
        tbl_row = 9
        ws_cm.cell(row=tbl_row, column=1, value="Detay Tablosu").font = Font(
            name="Segoe UI", bold=True, color=S["ACCENT"], size=11)
        tbl_row += 1
        cm_tbl_rows = []
        for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            cm = cm_all.get(sk)
            if cm is None:
                continue
            tn, fp, fn, tp = cm[0][0], cm[0][1], cm[1][0], cm[1][1]
            total = tn + fp + fn + tp
            cm_tbl_rows.append({
                "Split": sn, "TN": tn, "FP": fp, "FN": fn, "TP": tp,
                "Accuracy %": round((tn + tp) / total * 100, 2) if total else 0,
                "Precision %": round(tp / (tp + fp) * 100, 2) if (tp + fp) else 0,
                "Recall %": round(tp / (tp + fn) * 100, 2) if (tp + fn) else 0,
            })
        if cm_tbl_rows:
            df_cm = pd.DataFrame(cm_tbl_rows)
            _xl_write_df(ws_cm, df_cm, tbl_row, S, left_align_cols={1},
                         num_fmt_cols={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0",
                                       6: "0.00", 7: "0.00", 8: "0.00"})
            _xl_auto_width(ws_cm, 8, min_w=14)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 10-12 — Rating Dağılımları (26 / 10 / Decile) — tablo + grafik
    # ═══════════════════════════════════════════════════════════════════════════
    probas = tab_data.get("probabilities", {})
    y_true_data = tab_data.get("y_true", {})
    if probas and probas.get("train"):
        for method_name, thresholds, max_r in [
            ("Rating 26", _RATING_26_THRESHOLDS, 25),
            ("Rating 10", _RATING_10_THRESHOLDS, 10),
            ("Decile", None, 10),
        ]:
            all_split_rows = {}  # {split: [rows]}
            for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
                p = probas.get(sk)
                y = y_true_data.get(sk)
                if p is None or y is None:
                    continue
                p_arr = np.array(p)
                y_arr = np.array(y)
                if thresholds is not None:
                    ratings = _assign_rating_thresholds(p_arr, thresholds)
                else:
                    ratings = _assign_rating_decile(p_arr)
                overall_br = float(y_arr.sum()) / len(y_arr) if len(y_arr) > 0 else 0.0
                rating_counts = {}
                for r in set(ratings):
                    mask = ratings == r
                    rating_counts[int(r)] = (int(mask.sum()), int(y_arr[mask].sum()))
                rows = []
                for r in range(1, max_r + 1):
                    n, bads = rating_counts.get(r, (0, 0))
                    goods = n - bads
                    bad_rate = round(bads / n * 100, 2) if n > 0 else 0.0
                    row = {"Rating": r, "Count": n,
                           "Count%": round(n / len(p_arr) * 100, 2) if len(p_arr) > 0 else 0.0,
                           "Bad": bads, "Good": goods, "Bad Rate %": bad_rate}
                    if thresholds is None:
                        decile_br = bads / n if n > 0 else 0.0
                        row["Lift"] = round(decile_br / overall_br, 2) if overall_br > 0 else 0.0
                    rows.append(row)
                all_split_rows[sn] = rows

            if not all_split_rows:
                continue

            sheet_name = method_name[:31]
            ws_rt = wb.create_sheet(sheet_name)
            ws_rt.sheet_properties.tabColor = "0EA5E9"
            _xl_write_title(ws_rt, f"Rating Dağılımı — {method_name}", 8, S)

            current_row = 3
            for sn, rows in all_split_rows.items():
                # Split başlığı
                ws_rt.cell(row=current_row, column=1, value=sn).font = Font(
                    name="Segoe UI", bold=True, color=S["ACCENT"], size=11)
                current_row += 1

                df_rt = pd.DataFrame(rows)
                col_count = len(df_rt.columns)
                num_fmts = {3: "0.00", 6: "0.00"}
                if "Lift" in df_rt.columns:
                    num_fmts[7] = "0.00"
                end_r = _xl_write_df(ws_rt, df_rt, current_row, S,
                                     num_fmt_cols=num_fmts)

                # Bad Rate renklendirme
                br_col = list(df_rt.columns).index("Bad Rate %") + 1
                for r in range(current_row + 1, end_r + 1):
                    cell = ws_rt.cell(row=r, column=br_col)
                    if isinstance(cell.value, (int, float)):
                        if cell.value >= 10:
                            cell.font = S["red_font"]
                        elif cell.value >= 5:
                            cell.font = S["orange_font"]

                # Grafik: Count bar + Bad Rate line → Plotly PNG
                labels = [str(r.get("Rating", r.get("Decile", ""))) for r in rows]
                counts = [r["Count"] for r in rows]
                bad_rates = [r["Bad Rate %"] for r in rows]

                # Bar chart (Count)
                fig_count = _go.Figure(_go.Bar(
                    x=labels, y=counts,
                    marker_color="#93C5FD", text=counts, textposition="outside",
                    textfont=dict(size=9),
                ))
                fig_count.update_layout(
                    title=dict(text=f"{sn} — Count", font=dict(size=13)),
                    plot_bgcolor="#1F2937", paper_bgcolor="#111827",
                    font=dict(color="white", size=10),
                    xaxis=dict(title="Rating", tickangle=-45),
                    yaxis=dict(title="Count"),
                    margin=dict(l=60, r=20, t=40, b=60),
                )
                img_count = _fig_to_xl_image(fig_count, 560, 340)
                ws_rt.add_image(img_count, f"{get_column_letter(col_count + 2)}{current_row}")

                # Line chart (Bad Rate %)
                fig_br = _go.Figure(_go.Scatter(
                    x=labels, y=bad_rates, mode="lines+markers+text",
                    line=dict(color="#DC2626", width=2),
                    marker=dict(size=6, color="#DC2626"),
                    text=[f"{v:.1f}%" for v in bad_rates], textposition="top center",
                    textfont=dict(size=8, color="#FCA5A5"),
                ))
                fig_br.update_layout(
                    title=dict(text=f"{sn} — Bad Rate %", font=dict(size=13)),
                    plot_bgcolor="#1F2937", paper_bgcolor="#111827",
                    font=dict(color="white", size=10),
                    xaxis=dict(title="Rating", tickangle=-45),
                    yaxis=dict(title="Bad Rate %"),
                    margin=dict(l=60, r=20, t=40, b=60),
                )
                img_br = _fig_to_xl_image(fig_br, 560, 340)
                ws_rt.add_image(img_br, f"{get_column_letter(col_count + 12)}{current_row}")

                current_row = end_r + max_r + 4  # grafik yüksekliğine boşluk bırak

            _xl_auto_width(ws_rt, col_count, min_w=10)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET — HHI (Yoğunlaşma)
    # ═══════════════════════════════════════════════════════════════════════════
    if probas and probas.get("train"):
        import numpy as np
        ws_hhi = wb.create_sheet("HHI")
        ws_hhi.sheet_properties.tabColor = "8B5CF6"
        _xl_write_title(ws_hhi, "HHI — Yoğunlaşma (26 Segment)", 4, S)

        current_row = 3
        for sn, sk in [("Train", "train"), ("Test", "test"), ("OOT", "oot")]:
            p = probas.get(sk)
            if p is None:
                continue
            p_arr = np.array(p)
            if len(p_arr) == 0:
                continue

            rows, total_hhi = _calc_hhi_table(p_arr, _RATING_26_THRESHOLDS)

            if total_hhi < 0.06:
                yorum = "Dengeli"
            elif total_hhi < 0.10:
                yorum = "Orta"
            else:
                yorum = "Yoğun"

            # Split başlığı
            ws_hhi.cell(row=current_row, column=1,
                        value=f"{sn}  ·  HHI = {total_hhi:.6f}  ·  {yorum}").font = Font(
                name="Segoe UI", bold=True, color=S["ACCENT"], size=11)
            current_row += 1

            df_hhi = pd.DataFrame(rows)
            end_r = _xl_write_df(ws_hhi, df_hhi, current_row, S,
                                 left_align_cols={1},
                                 num_fmt_cols={3: "0.00", 4: "0.000000"})
            # TOPLAM satırı kalın
            for c in range(1, 5):
                cell = ws_hhi.cell(row=end_r, column=c)
                cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor="1E293B")
            current_row = end_r + 2

        _xl_auto_width(ws_hhi, 4, min_w=12)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET — Korelasyon
    # ═══════════════════════════════════════════════════════════════════════════
    corr_dict = results.get("raw_corr") if tab_key == "raw" else results.get("corr")
    if corr_dict:
        vars_ = list(corr_dict.keys())
        n_vars = len(vars_)
        if n_vars >= 2:
            ws_cr = wb.create_sheet("Korelasyon")
            ws_cr.sheet_properties.tabColor = "F43F5E"
            _xl_write_title(ws_cr, f"Korelasyon Matrisi ({n_vars} değişken)", n_vars + 1, S)

            # Header satırı
            ws_cr.cell(row=3, column=1, value="")
            for ci, v in enumerate(vars_, 2):
                ws_cr.cell(row=3, column=ci, value=v)
            _xl_style_header(ws_cr, 3, n_vars + 1, S)

            # Veri ve koşullu renklendirme
            from openpyxl.styles import numbers as xl_numbers

            def _corr_fill(val):
                """Korelasyon değerine göre fill rengi — Reds tonu."""
                av = abs(val)
                if av >= 0.8:
                    return PatternFill("solid", fgColor="FCA5A5")  # koyu kırmızı
                if av >= 0.6:
                    return PatternFill("solid", fgColor="FECACA")
                if av >= 0.4:
                    return PatternFill("solid", fgColor="FEE2E2")
                if av >= 0.2:
                    return PatternFill("solid", fgColor="FEF2F2")
                return S["odd_fill"]

            for ri, rv in enumerate(vars_):
                r = 4 + ri
                ws_cr.cell(row=r, column=1, value=rv).font = S["bold_font"]
                ws_cr.cell(row=r, column=1).fill = S["hdr_fill"]
                ws_cr.cell(row=r, column=1).font = S["hdr_font"]
                ws_cr.cell(row=r, column=1).border = S["border"]
                ws_cr.cell(row=r, column=1).alignment = S["cell_align_l"]
                for ci, cv in enumerate(vars_):
                    val = corr_dict[rv][cv]
                    cell = ws_cr.cell(row=r, column=2 + ci, value=round(val, 4))
                    cell.number_format = "0.0000"
                    cell.font = S["cell_font"]
                    cell.alignment = S["cell_align"]
                    cell.border = S["border"]
                    cell.fill = _corr_fill(val)
                    # Yüksek korelasyonları kalın kırmızı yap
                    if abs(val) >= 0.7 and rv != cv:
                        cell.font = S["red_font"]

            _xl_auto_width(ws_cr, n_vars + 1, min_w=12, max_w=20)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 14 — WoE Dağılımı (sadece WoE sekmesi)
    # ═══════════════════════════════════════════════════════════════════════════
    if tab_key == "woe":
        woe_dist = results.get("woe_dist")
        if woe_dist:
            ws_woe = wb.create_sheet("WoE Dağılımı")
            ws_woe.sheet_properties.tabColor = "A78BFA"
            _xl_write_title(ws_woe, "WoE Bin Dağılımları", 10, S)

            current_row = 3
            for var_name, info in woe_dist.items():
                train_rows = info.get("train_table")
                if not train_rows:
                    continue
                mono_train = info.get("monoton", "–")
                mono_test  = info.get("monoton_test", "–")
                mono_oot   = info.get("monoton_oot", "–")
                iv_train = info.get("iv_train", 0)
                iv_test  = info.get("iv_test")
                iv_oot   = info.get("iv_oot")

                def _xl_mono_color(m):
                    if "Artan" in m or "Azalan" in m:
                        return S["GREEN"]
                    if "Değil" in m or "\u2717" in m:
                        return S["RED"]
                    return "6B7280"

                # Değişken başlığı
                ws_woe.cell(row=current_row, column=1, value=var_name).font = Font(
                    name="Segoe UI", bold=True, color="FFFFFF", size=11)
                ws_woe.cell(row=current_row, column=1).fill = PatternFill(
                    "solid", fgColor="374151")

                # IV + monotonluk: her split yan yana
                iv_text = f"IV (Train): {iv_train}  [{mono_train}]"
                if iv_test is not None:
                    iv_text += f"  |  IV (Test): {iv_test}  [{mono_test}]"
                if iv_oot is not None:
                    iv_text += f"  |  IV (OOT): {iv_oot}  [{mono_oot}]"
                ws_woe.cell(row=current_row, column=3, value=iv_text).font = Font(
                    name="Segoe UI", color="8B5CF6", size=10)

                current_row += 1

                # Train tablosu
                df_train = pd.DataFrame(train_rows)
                ws_woe.cell(row=current_row, column=1, value="Train").font = Font(
                    name="Segoe UI", bold=True, color=S["ACCENT"], size=10)
                current_row += 1
                end_r = _xl_write_df(ws_woe, df_train, current_row, S, left_align_cols={1})
                current_row = end_r + 2

                # Test tablosu
                test_rows = info.get("test_table")
                if test_rows:
                    ws_woe.cell(row=current_row, column=1, value="Test").font = Font(
                        name="Segoe UI", bold=True, color=S["ACCENT"], size=10)
                    current_row += 1
                    end_r = _xl_write_df(ws_woe, pd.DataFrame(test_rows), current_row, S, left_align_cols={1})
                    current_row = end_r + 2

                # OOT tablosu
                oot_rows = info.get("oot_table")
                if oot_rows:
                    ws_woe.cell(row=current_row, column=1, value="OOT").font = Font(
                        name="Segoe UI", bold=True, color=S["ACCENT"], size=10)
                    current_row += 1
                    end_r = _xl_write_df(ws_woe, pd.DataFrame(oot_rows), current_row, S, left_align_cols={1})
                    current_row = end_r + 2

                current_row += 1  # değişkenler arası boşluk

            _xl_auto_width(ws_woe, 10)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 15 — Describe
    # ═══════════════════════════════════════════════════════════════════════════
    describe_data = results.get("describe_data")
    if describe_data:
        ws_d = wb.create_sheet("Describe")
        ws_d.sheet_properties.tabColor = "64748B"
        _xl_write_title(ws_d, "Değişken İstatistikleri (Describe)", len(describe_data[0]), S)
        df_desc = pd.DataFrame(describe_data)
        end_r = _xl_write_df(ws_d, df_desc, 3, S, left_align_cols={1})
        _xl_auto_width(ws_d, len(df_desc.columns))

    # ═══════════════════════════════════════════════════════════════════════════
    # Dosyayı profil klasörüne kaydet
    # ═══════════════════════════════════════════════════════════════════════════
    profile_dir = Path(__file__).parent.parent / "profiles" / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    xl_path = profile_dir / f"{fname}.xlsx"
    wb.save(xl_path)
    return dbc.Alert(f"✓ Excel kaydedildi → {xl_path.name}",
                     color="success", style=_A)
