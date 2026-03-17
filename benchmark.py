"""
EDA Laboratuvarı — Benchmark Testi
====================================
5.7M satır × 37 kolon veri seti üzerinde modül bazında süre ve RAM ölçümü.
Sonuçlar benchmark_results.xlsx dosyasına kaydedilir.
"""

import os
import sys
import gc
import re
import glob
import time
import tracemalloc
import warnings
import platform
import psutil

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8") if hasattr(sys.stdout, "reconfigure") else None

# ─────────────────────────────────────────────────────────────────────────────
# AYARLAR — Gerekirse düzenle
# ─────────────────────────────────────────────────────────────────────────────
DATA_ROOT   = r"C:\Users\pc\OneDrive\Desktop\inghubstr-ds-case-study-coderspace\inghubstr-ds-case-study-coderspace\data"
TARGET_COL  = "has_cc"          # Binary hedef kolon
DATE_COL    = "month_date"
SEGMENT_COL = "segment"
CUST_COL    = "cust_id"
OUTPUT_FILE = "benchmark_results.xlsx"

SEGMENT_RATIOS = [1.0, 0.5, 0.1]   # Tam veri / %50 / %10

# ─────────────────────────────────────────────────────────────────────────────
# YARDIMCI FONKSİYONLAR
# ─────────────────────────────────────────────────────────────────────────────

def mb(bytes_val: int) -> float:
    return round(bytes_val / 1024 / 1024, 1)

def mem_now() -> float:
    process = psutil.Process(os.getpid())
    return round(process.memory_info().rss / 1024 / 1024, 1)

def run_test(label: str, fn, *args, **kwargs):
    """
    Fonksiyonu çalıştırır, süre ve peak RAM delta ölçer.
    Returns: (label, duration_sec, peak_ram_delta_mb, status, notes)
    """
    print(f"  ▶ {label}...", end=" ", flush=True)
    gc.collect()
    ram_before = mem_now()
    tracemalloc.start()
    t0 = time.perf_counter()
    status = "OK"
    notes  = ""
    result = None
    try:
        result = fn(*args, **kwargs)
    except Exception as e:
        status = "ERROR"
        notes  = str(e)[:120]
    elapsed = round(time.perf_counter() - t0, 2)
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    ram_after = mem_now()
    peak_delta = round(ram_after - ram_before, 1)
    print(f"{elapsed}s | RAM +{peak_delta}MB | {status}")
    return {
        "Test": label,
        "Süre (sn)": elapsed,
        "Peak RAM Delta (MB)": peak_delta,
        "Peak RAM (tracemalloc MB)": mb(peak),
        "Durum": status,
        "Notlar": notes,
    }, result


# ─────────────────────────────────────────────────────────────────────────────
# VERİ YÜKLEME
# ─────────────────────────────────────────────────────────────────────────────

def get_date_from_filename(filename):
    match = re.search(r'(\d{4})[_-]?(\d{2})', filename)
    if match:
        year, month = match.groups()
        return pd.to_datetime(f"{year}-{month}-01")
    return None

def read_parquet_robust(path):
    try:
        return pd.read_parquet(path)
    except Exception:
        try:
            return pd.read_parquet(path, engine='fastparquet')
        except Exception as e:
            print(f"  HATA: {os.path.basename(path)}: {e}")
            return None

def read_monthly_data(folder_path):
    files = glob.glob(os.path.join(folder_path, "*.parquet"))
    df_list = []
    for f in files:
        temp_df = read_parquet_robust(f)
        if temp_df is not None:
            date_val = get_date_from_filename(os.path.basename(f))
            if date_val:
                temp_df['month_date'] = date_val
                df_list.append(temp_df)
    return pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()

def load_all_data():
    paths = {
        'demo':        os.path.join(DATA_ROOT, 'demographic.csv'),
        'activity':    os.path.join(DATA_ROOT, 'activity'),
        'ownership':   os.path.join(DATA_ROOT, 'ownership'),
        'transaction': os.path.join(DATA_ROOT, 'transaction'),
    }
    df_demo        = pd.read_csv(paths['demo'])
    df_activity    = read_monthly_data(paths['activity'])
    df_ownership   = read_monthly_data(paths['ownership'])
    df_transaction = read_monthly_data(paths['transaction'])

    df = df_transaction.copy()
    if not df_activity.empty:
        df = df.merge(df_activity,  on=['cust_id', 'month_date'], how='left')
    if not df_ownership.empty:
        df = df.merge(df_ownership, on=['cust_id', 'month_date'], how='left')
    if not df_demo.empty:
        df = df.merge(df_demo, on='cust_id', how='left')
    return df


# ─────────────────────────────────────────────────────────────────────────────
# OUTLIER YARDIMCILARI (modules/outlier yoksa inline)
# ─────────────────────────────────────────────────────────────────────────────

def _outlier_iqr(df: pd.DataFrame, cols: list, k: float = 1.5) -> pd.Series:
    mask = pd.Series(False, index=df.index)
    for c in cols:
        s  = df[c].dropna()
        q1, q3 = s.quantile(0.25), s.quantile(0.75)
        iqr = q3 - q1
        lo, hi = q1 - k * iqr, q3 + k * iqr
        mask |= ((df[c] < lo) | (df[c] > hi))
    return mask

def _outlier_zscore(df: pd.DataFrame, cols: list, thresh: float = 3.0) -> pd.Series:
    mask = pd.Series(False, index=df.index)
    for c in cols:
        mu  = df[c].mean()
        std = df[c].std()
        if std > 0:
            mask |= ((df[c] - mu).abs() / std > thresh)
    return mask

def _outlier_per_col(df: pd.DataFrame, cols: list, method: str, param: float) -> pd.DataFrame:
    """Her kolon için outlier mask'i döndürür (bool DataFrame)."""
    result = pd.DataFrame(False, index=df.index, columns=cols)
    for c in cols:
        s = df[c]
        if method == "iqr":
            q1, q3 = s.quantile(0.25), s.quantile(0.75)
            iqr = q3 - q1
            result[c] = (s < q1 - param * iqr) | (s > q3 + param * iqr)
        else:
            mu, std = s.mean(), s.std()
            if std > 0:
                result[c] = (s - mu).abs() / std > param
    return result


# ─────────────────────────────────────────────────────────────────────────────
# ANA BENCHMARK
# ─────────────────────────────────────────────────────────────────────────────

def main():
    results_load   = []
    results_module = []
    results_seg    = []
    results_model  = []
    results_shap   = []

    sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
    from modules import profiling, target_analysis, screening, correlation, deep_dive

    # ── Sistem Bilgisi ──────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  EDA Laboratuvari -- Benchmark Testi")
    print("=" * 60)
    print(f"  Platform : {platform.platform()}")
    print(f"  Python   : {sys.version.split()[0]}")
    cpu_info = f"{psutil.cpu_count(logical=True)} mantıksal çekirdek"
    ram_total = round(psutil.virtual_memory().total / 1024**3, 1)
    print(f"  CPU      : {cpu_info}")
    print(f"  RAM      : {ram_total} GB toplam | {mem_now()} MB kullanımda")
    print(f"  Pandas   : {pd.__version__}")
    import sklearn; print(f"  sklearn  : {sklearn.__version__}")
    try:
        import lightgbm as lgb; print(f"  lightgbm : {lgb.__version__}")
    except: print("  lightgbm : YOK")
    try:
        import xgboost as xgb; print(f"  xgboost  : {xgb.__version__}")
    except: print("  xgboost  : YOK")
    try:
        import shap; print(f"  shap     : {shap.__version__}")
    except: print("  shap     : YOK")

    # ══════════════════════════════════════════════════════════════════════
    # FAZ 1 — VERİ YÜKLEME
    # ══════════════════════════════════════════════════════════════════════
    print("\n[FAZ 1] Veri Yükleme")
    row, df_full = run_test("Veri Yükleme (Parquet + CSV + Merge)", load_all_data)
    results_load.append(row)

    if df_full is None or df_full.empty:
        print("HATA: Veri yüklenemedi, test sonlandırılıyor.")
        return

    print(f"  → Veri boyutu: {df_full.shape[0]:,} satır × {df_full.shape[1]} kolon")
    print(f"  → Kolonlar: {df_full.columns.tolist()}")
    print(f"  → RAM sonrası: {mem_now()} MB")

    num_cols = [c for c in df_full.select_dtypes(include=[np.number]).columns
                if c not in [TARGET_COL, CUST_COL]]
    cat_cols = [c for c in df_full.select_dtypes(exclude=[np.number]).columns
                if c not in [DATE_COL]]

    # Target sütunu kontrol
    has_target = TARGET_COL in df_full.columns
    print(f"  → Target '{TARGET_COL}': {'VAR' if has_target else 'YOK — IV/WOE testleri atlanacak'}")

    # ══════════════════════════════════════════════════════════════════════
    # FAZ 2 — MODÜL TESTLERİ (tam veri)
    # ══════════════════════════════════════════════════════════════════════
    print("\n[FAZ 2] Modül Testleri — Tam Veri")

    # Profiling
    row, _ = run_test("Profiling — compute_profile()", profiling.compute_profile, df_full)
    results_module.append(row)

    # Target & IV
    if has_target:
        row, _ = run_test(
            "Target — compute_target_stats()",
            target_analysis.compute_target_stats, df_full, TARGET_COL
        )
        results_module.append(row)

        row, _ = run_test(
            "IV Ranking — compute_iv_ranking_optimal() [tüm num kolonlar]",
            deep_dive.compute_iv_ranking_optimal, df_full, TARGET_COL, num_cols
        )
        results_module.append(row)

    # Screening
    row, _ = run_test(
        "Screening — screen_columns()",
        screening.screen_columns, df_full, TARGET_COL if has_target else num_cols[0],
        DATE_COL, SEGMENT_COL
    )
    results_module.append(row)

    # Korelasyon
    corr_cols = num_cols[:30]   # max 30 kolon
    row, _ = run_test(
        f"Korelasyon — compute_correlation_matrix() [{len(corr_cols)} kolon]",
        correlation.compute_correlation_matrix, df_full, corr_cols
    )
    results_module.append(row)

    # VIF
    row, _ = run_test(
        f"VIF — compute_vif() [{len(corr_cols)} kolon]",
        correlation.compute_vif, df_full, corr_cols
    )
    results_module.append(row)

    # WOE Detail (tek kolon)
    if has_target and num_cols:
        woe_col = num_cols[0]
        row, _ = run_test(
            f"WOE Detail — get_woe_detail() [1 kolon: {woe_col}]",
            deep_dive.get_woe_detail, df_full, woe_col, TARGET_COL
        )
        results_module.append(row)

        cutoff = str(df_full[DATE_COL].dropna().sort_values().iloc[len(df_full) // 2])[:10]
        row, _ = run_test(
            f"PSI — compute_psi() [1 kolon: {woe_col}, cutoff: {cutoff}]",
            deep_dive.compute_psi, df_full, woe_col, TARGET_COL, DATE_COL, cutoff
        )
        results_module.append(row)

        row, _ = run_test(
            f"Variable Stats — get_variable_stats() [1 kolon: {woe_col}]",
            deep_dive.get_variable_stats, df_full, woe_col, TARGET_COL
        )
        results_module.append(row)

    # Target over time
    row, _ = run_test(
        "Target Over Time — compute_target_over_time()",
        target_analysis.compute_target_over_time, df_full, TARGET_COL, DATE_COL
    )
    results_module.append(row)

    # High corr pairs
    def _find_high_corr():
        corr_mat = correlation.compute_correlation_matrix(df_full, corr_cols)
        return correlation.find_high_corr_pairs(corr_mat, threshold=0.7)

    row, _ = run_test(
        f"High Corr Pairs — find_high_corr_pairs() [{len(corr_cols)} kolon, threshold=0.7]",
        _find_high_corr
    )
    results_module.append(row)

    # ── İstatistiksel Testler ─────────────────────────────────────────────
    from scipy import stats as scipy_stats

    # Chi-Square: gender vs has_cc
    def chi_square_test():
        s1 = df_full["gender"].fillna("(bos)").astype(str)
        s2 = df_full[TARGET_COL].fillna(0).astype(str)
        ctab = pd.crosstab(s1, s2)
        chi2, p, dof, _ = scipy_stats.chi2_contingency(ctab)
        return chi2, p, dof

    row, _ = run_test("Chi-Square — gender vs has_cc (5.7M satir)", chi_square_test)
    results_module.append(row)

    # ANOVA: cc_trx_amt ~ has_cc (app ile ayni: grup basina max 200k ornek)
    def anova_test():
        col_data = df_full[["cc_trx_amt", TARGET_COL]].dropna()
        groups = col_data.groupby(TARGET_COL)["cc_trx_amt"]
        MAX_PER_GROUP = 200_000
        rng = np.random.default_rng(42)
        sampled = [
            g.values if len(g.values) <= MAX_PER_GROUP
            else rng.choice(g.values, MAX_PER_GROUP, replace=False)
            for _, g in groups
        ]
        return scipy_stats.f_oneway(*sampled)

    row, _ = run_test("ANOVA — cc_trx_amt ~ has_cc (max 200k/grup ornekleme)", anova_test)
    results_module.append(row)

    # KS Test: cc_trx_amt good vs bad
    def ks_test():
        col_data = df_full[["cc_trx_amt", TARGET_COL]].dropna()
        good = col_data[col_data[TARGET_COL] == 0]["cc_trx_amt"].values
        bad  = col_data[col_data[TARGET_COL] == 1]["cc_trx_amt"].values
        return scipy_stats.ks_2samp(good, bad)

    row, _ = run_test("KS Test — cc_trx_amt (good vs bad, tam veri)", ks_test)
    results_module.append(row)

    # Outlier — IQR
    out_cols = num_cols[:20]
    row, _ = run_test(
        f"Outlier IQR 1.5 — [{len(out_cols)} kolon]",
        _outlier_iqr, df_full, out_cols, 1.5
    )
    results_module.append(row)

    row, _ = run_test(
        f"Outlier IQR 3.0 — [{len(out_cols)} kolon]",
        _outlier_iqr, df_full, out_cols, 3.0
    )
    results_module.append(row)

    # Outlier — Z-Score
    row, _ = run_test(
        f"Outlier Z-Score 3.0 — [{len(out_cols)} kolon]",
        _outlier_zscore, df_full, out_cols, 3.0
    )
    results_module.append(row)

    # Müşteri bazında outlier sayısı
    def customer_outlier_table(df, cols, cust_col, method, param):
        bool_df = _outlier_per_col(df, cols, method, param)
        df_res = df[[cust_col]].copy()
        df_res["outlier_col_count"] = bool_df.sum(axis=1)
        return df_res.groupby(cust_col)["outlier_col_count"].max().reset_index()

    row, _ = run_test(
        f"Müşteri Outlier Tablosu (IQR 1.5) [{len(out_cols)} kolon]",
        customer_outlier_table, df_full, out_cols, CUST_COL, "iqr", 1.5
    )
    results_module.append(row)

    # ══════════════════════════════════════════════════════════════════════
    # FAZ 3 — SEGMENT FİLTRE ETKİSİ
    # ══════════════════════════════════════════════════════════════════════
    print("\n[FAZ 3] Segment Filtresi Etkisi")

    for ratio in SEGMENT_RATIOS:
        n_sample = int(len(df_full) * ratio)
        df_seg = df_full.sample(n=n_sample, random_state=42) if ratio < 1.0 else df_full
        label_prefix = f"Segment %{int(ratio*100)} ({n_sample:,} satır)"

        row, _ = run_test(
            f"{label_prefix} — Profiling",
            profiling.compute_profile, df_seg
        )
        row["Test"] = f"{label_prefix} | Profiling"
        results_seg.append(row)

        if has_target:
            row, _ = run_test(
                f"{label_prefix} — IV Ranking",
                deep_dive.compute_iv_ranking_optimal, df_seg, TARGET_COL, num_cols
            )
            row["Test"] = f"{label_prefix} | IV Ranking"
            results_seg.append(row)

        row, _ = run_test(
            f"{label_prefix} — Outlier IQR",
            _outlier_iqr, df_seg, out_cols, 1.5
        )
        row["Test"] = f"{label_prefix} | Outlier IQR"
        results_seg.append(row)

    # ══════════════════════════════════════════════════════════════════════
    # FAZ 4 — MODEL & SHAP
    # ══════════════════════════════════════════════════════════════════════
    print("\n[FAZ 4] Model Eğitimi & SHAP")

    if not has_target:
        print("  Target kolonu yok — model testleri atlanıyor.")
    else:
        from sklearn.model_selection import train_test_split
        from sklearn.linear_model import LogisticRegression
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.preprocessing import StandardScaler

        X = df_full[num_cols].fillna(df_full[num_cols].median())
        y = df_full[TARGET_COL].fillna(0).astype(int)
        X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.3, random_state=42)

        print(f"  Train: {X_tr.shape[0]:,} | Test: {X_te.shape[0]:,}")

        # Logistic Regression
        def fit_lr():
            sc = StandardScaler()
            Xsc = sc.fit_transform(X_tr)
            clf = LogisticRegression(max_iter=200, solver='saga', n_jobs=-1)
            clf.fit(Xsc, y_tr)
            return clf, sc

        row, lr_res = run_test("Model — Logistic Regression (fit)", fit_lr)
        results_model.append(row)

        # LightGBM
        try:
            import lightgbm as lgb
            def fit_lgbm():
                clf = lgb.LGBMClassifier(n_estimators=100, n_jobs=-1, verbose=-1)
                clf.fit(X_tr, y_tr)
                return clf
            row, lgbm_model = run_test("Model — LightGBM (fit)", fit_lgbm)
            results_model.append(row)
        except ImportError:
            lgbm_model = None
            print("  LightGBM yüklü değil — atlanıyor.")

        # XGBoost
        try:
            import xgboost as xgb
            def fit_xgb():
                clf = xgb.XGBClassifier(n_estimators=100, n_jobs=-1,
                                         use_label_encoder=False, eval_metric='logloss',
                                         verbosity=0)
                clf.fit(X_tr, y_tr)
                return clf
            row, xgb_model = run_test("Model — XGBoost (fit)", fit_xgb)
            results_model.append(row)
        except ImportError:
            xgb_model = None
            print("  XGBoost yüklü değil — atlanıyor.")

        # Random Forest
        def fit_rf():
            clf = RandomForestClassifier(n_estimators=100, n_jobs=-1, random_state=42)
            clf.fit(X_tr, y_tr)
            return clf
        row, _ = run_test("Model — Random Forest (fit)", fit_rf)
        results_model.append(row)

        # SHAP — tüm test seti
        try:
            import shap

            if lgbm_model is not None:
                def shap_lgbm_full():
                    exp = shap.TreeExplainer(lgbm_model)
                    return exp.shap_values(X_te)
                row, _ = run_test(f"SHAP — LightGBM (tüm test: {len(X_te):,} satır)", shap_lgbm_full)
                results_shap.append(row)

            if xgb_model is not None:
                def shap_xgb_full():
                    exp = shap.TreeExplainer(xgb_model)
                    return exp.shap_values(X_te)
                row, _ = run_test(f"SHAP — XGBoost (tüm test: {len(X_te):,} satır)", shap_xgb_full)
                results_shap.append(row)

            # RF SHAP atlandi — 1.7M satir * 100 agac, saat alir
            results_shap.append({
                "Test": f"SHAP — Random Forest (tüm test: {len(X_te):,} satır)",
                "Süre (sn)": "ATLANDI",
                "Peak RAM Delta (MB)": "-",
                "Peak RAM (tracemalloc MB)": "-",
                "Durum": "SKIP",
                "Notlar": "RF SHAP 1.7M satirda saat mertebesinde suruyor, kasitli atlanmistir",
            })

        except ImportError:
            print("  SHAP yüklü değil — atlanıyor.")
        except Exception as e:
            print(f"  SHAP hatası: {e}")

    # ══════════════════════════════════════════════════════════════════════
    # EXCEL RAPORU
    # ══════════════════════════════════════════════════════════════════════
    print(f"\n[RAPOR] Excel'e yazılıyor → {OUTPUT_FILE}")

    env_data = {
        "Parametre": [
            "Tarih", "Platform", "Python", "Pandas",
            "Toplam RAM (GB)", "CPU Çekirdeği",
            "Veri Boyutu (satır)", "Veri Boyutu (kolon)",
            "Target Kolon",
        ],
        "Değer": [
            pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
            platform.platform(),
            sys.version.split()[0],
            pd.__version__,
            ram_total,
            psutil.cpu_count(logical=True),
            f"{df_full.shape[0]:,}",
            df_full.shape[1],
            TARGET_COL if has_target else "YOK",
        ]
    }

    kabul_data = {
        "Operasyon": [
            "Veri Yükleme (5M+ satır)",
            "Profiling",
            "IV Ranking",
            "Outlier Tara",
            "Korelasyon Matrisi",
            "VIF",
            "Model (LightGBM)",
            "Model (RF)",
            "SHAP (tüm test seti)",
        ],
        "Hedef (sn)": [30, 15, 20, 25, 20, 30, 60, 120, 120],
        "Not": [
            "Parquet + merge",
            "Tüm kolonlar",
            "Tüm numerik kolonlar",
            "IQR 1.5, 20 kolon",
            "30 kolon",
            "30 kolon",
            "100 estimator",
            "100 estimator",
            "Tüm test seti",
        ]
    }

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        pd.DataFrame(env_data).to_excel(writer, sheet_name="Ortam", index=False)
        pd.DataFrame(results_load).to_excel(writer, sheet_name="Faz1_VeriYükleme", index=False)
        pd.DataFrame(results_module).to_excel(writer, sheet_name="Faz2_Modüller", index=False)
        pd.DataFrame(results_seg).to_excel(writer, sheet_name="Faz3_SegmentFiltre", index=False)
        pd.DataFrame(results_model).to_excel(writer, sheet_name="Faz4_Model", index=False)
        pd.DataFrame(results_shap).to_excel(writer, sheet_name="Faz4_SHAP", index=False)
        pd.DataFrame(kabul_data).to_excel(writer, sheet_name="KabulKriterleri", index=False)

        # Tüm sonuçlar tek sayfada
        all_rows = results_load + results_module + results_seg + results_model + results_shap
        df_all = pd.DataFrame(all_rows)
        df_all.to_excel(writer, sheet_name="Özet_Tümü", index=False)

        # Koşullu biçimlendirme — süre > 60sn kırmızı
        from openpyxl.styles import PatternFill, Font
        red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        orange_fill= PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

        for sheet_name in ["Faz2_Modüller", "Faz4_Model", "Faz4_SHAP", "Özet_Tümü"]:
            if sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2 and cell.value is not None:  # Süre sütunu
                        try:
                            v = float(cell.value)
                            if v > 120:
                                cell.fill = red_fill
                            elif v > 30:
                                cell.fill = orange_fill
                            else:
                                cell.fill = green_fill
                        except: pass
                    if cell.column == 5:  # Durum sütunu
                        if cell.value == "ERROR":
                            cell.fill = red_fill
                            cell.font = Font(bold=True)

    print(f"\nTamamlandi -> {os.path.abspath(OUTPUT_FILE)}")
    print(f"  Toplam test sayısı: {len(results_load) + len(results_module) + len(results_seg) + len(results_model) + len(results_shap)}")


if __name__ == "__main__":
    main()
